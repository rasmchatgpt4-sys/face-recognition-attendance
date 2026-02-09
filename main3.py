# main.py
"""
Face Recognition Based Attendance System (ORB fallback; no cv2.face required)

- Uses ORB (pure OpenCV) for feature-based face recognition.
- Captures training images into TrainingImage/ as before.
- Builds ORB index (descriptors per label) and caches it to Models/orb_index.npz.
- Recognition compares descriptors from camera faces to per-label descriptors and picks best match.
- Includes voice assistant (Alan) with wake-word + manual query (Start Query / Stop Query).
- Attendance saves CSV session files and updates Excel registers (if openpyxl available).
- All UI text uses black foreground for readability.
- Inline comments added to explain complex parts.

Run: python main.py
Requires: OpenCV (opencv-python or opencv-contrib-python), numpy, pandas, Pillow, speech_recognition, pyttsx3.
Optional: openpyxl for registers, reportlab for PDF reports, scikit-learn/joblib for ML features.
"""

import os
import re
import threading
import time
from datetime import datetime
import difflib
import glob
import json

import cv2
import numpy as np
from PIL import Image
import pandas as pd

import tkinter as tk
from tkinter import messagebox, filedialog, ttk

# speech & tts
import speech_recognition as sr
import pyttsx3

# excel writer (optional)
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment
except Exception:
    Workbook = None
    load_workbook = None
    get_column_letter = None
    Alignment = None

# ------------------- PATHS & DIRECTORIES -------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

CASCADE_PATH = os.path.join(BASE_DIR, "haarcascade_frontalface_default.xml")

TRAINING_IMAGE_DIR = os.path.join(BASE_DIR, "TrainingImage")
os.makedirs(TRAINING_IMAGE_DIR, exist_ok=True)

MODEL_DIR = os.path.join(BASE_DIR, "Models")
os.makedirs(MODEL_DIR, exist_ok=True)
ORB_INDEX_PATH = os.path.join(MODEL_DIR, "orb_index.npz")  # cached ORB descriptors

TRAINER_DIR = os.path.join(BASE_DIR, "TrainingImageLabel")
os.makedirs(TRAINER_DIR, exist_ok=True)

STUDENT_CSV = os.path.join(BASE_DIR, "StudentDetails.csv")
ATTENDANCE_DIR = os.path.join(BASE_DIR, "Attendance")
os.makedirs(ATTENDANCE_DIR, exist_ok=True)

ATTENDANCE_REGISTERS_DIR = os.path.join(BASE_DIR, "AttendanceRegisters")
os.makedirs(ATTENDANCE_REGISTERS_DIR, exist_ok=True)

REPORT_DIR = os.path.join(BASE_DIR, "Reports")
os.makedirs(REPORT_DIR, exist_ok=True)

# ------------------- UI THEME (text in black) -------------------
BG_MAIN = "#f0f4ff"
BG_CARD = "#ffffff"
BUTTON_PRIMARY = "#667eea"
BUTTON_FG = "#000000"
ENTRY_BG = "#ffffff"
ENTRY_FG = "#000000"
FG_MAIN = "#000000"

TITLE_FONT = ("Segoe UI", 18, "bold")
LABEL_FONT = ("Segoe UI", 11)
BUTTON_FONT = ("Segoe UI", 10, "bold")
SMALL_FONT = ("Segoe UI", 9)

# ------------------- FACE/ORB SETTINGS -------------------
FACE_SIZE = (200, 200)  # normalized face size for descriptor extraction
ORB_N_FEATURES = 1000   # number of ORB features to detect
ORB_MATCH_THRESHOLD = 18   # minimum number of "good" matches to consider a label (tune)
ORB_GOOD_DISTANCE = 60     # BFMatcher distance threshold for "good" match (lower = stricter)
ORB_RATIO_MIN = 0.10       # alternative ratio of goodMatches/num_kp to accept (tune)

# ------------------- Semester/Subjects (trimmed copy) -------------------
SEMESTER_OPTIONS = ["3rd Semester", "5th Semester", "7th Semester"]
SUBJECTS_BY_SEM = {
    "3rd Semester": ["DMGT", "DSC (IPCC)", "OOPS (IPCC)", "DSDO", "PAI", "SC & R"],
    "5th Semester": ["DS", "PML", "DBMS", "HCI"],
    "7th Semester": ["BDA", "GEN AI", "CN"],
}

# ------------------- AUTH / FACULTY (kept simple) -------------------
FACULTY_INFO = {
    "kaveri_faculty": {"display_name": "Prof. Kaveri K", "sem_subjects": {"3rd Semester": ["PAI", "SC & R"]}, "sections": {"3rd Semester": "III"}},
    "sowmya_faculty": {"display_name": "Prof. Sowmya G", "sem_subjects": {"3rd Semester": ["DMGT"]}, "sections": {"3rd Semester": "III"}},
}
AUTH = {
    "faculty": {"kaveri": {"password": "123", "faculty_key": "kaveri_faculty"}, "sowmya": {"password": "123", "faculty_key": "sowmya_faculty"}},
    "hod": {"hod": {"password": "123", "name": "HOD-AIML"}},
    "principal": {"principal": {"password": "123", "name": "Principal"}},
}

current_role = None
current_faculty_name = None
current_faculty_sem_subjects = {}
current_faculty_sections = {}

# ------------------- CSV helpers -------------------
def ensure_student_csv():
    if not os.path.exists(STUDENT_CSV):
        df = pd.DataFrame(columns=["Label", "USN", "Name", "Semester"])
        df.to_csv(STUDENT_CSV, index=False)

ensure_student_csv()

def _load_student_df():
    ensure_student_csv()
    try:
        return pd.read_csv(STUDENT_CSV, dtype=str)
    except Exception:
        return pd.DataFrame(columns=["Label", "USN", "Name", "Semester"])

def save_student_details(usn: str, student_name: str, semester: str):
    usn = usn.strip()
    student_name = student_name.strip()
    semester = semester.strip()
    if not usn or not student_name:
        messagebox.showerror("Error", "USN and Name are required.")
        return False
    df = _load_student_df()
    if ((df["USN"].astype(str).str.lower()) == usn.lower()).any():
        messagebox.showinfo("Info", f"USN {usn} already exists.")
        return False
    if df.empty:
        next_label = 1
    else:
        try:
            labels = df["Label"].dropna().astype(int).tolist()
            next_label = max(labels) + 1 if labels else 1
        except Exception:
            next_label = 1
    new_row = {"Label": str(next_label), "USN": usn, "Name": student_name, "Semester": semester}
    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    tmp = STUDENT_CSV + ".tmp"
    df.to_csv(tmp, index=False)
    try:
        os.replace(tmp, STUDENT_CSV)
    except Exception:
        df.to_csv(STUDENT_CSV, index=False)
    set_status(f"Saved student {usn} - {student_name} (label {next_label}).")
    # Invalidate ORB cache so next recognition rebuilds index
    try:
        if os.path.exists(ORB_INDEX_PATH):
            os.remove(ORB_INDEX_PATH)
    except Exception:
        pass
    return True

def remove_student_by_usn(usn: str):
    usn = usn.strip()
    if not usn:
        messagebox.showerror("Error", "Enter USN")
        return False
    df = _load_student_df()
    mask = df["USN"].astype(str).str.lower() == usn.lower()
    if not mask.any():
        messagebox.showinfo("Info", f"No student with USN {usn}")
        return False
    matched = df[mask].iloc[0]
    label = str(matched["Label"])
    name = matched["Name"]
    if not messagebox.askyesno("Confirm", f"Remove {name} ({usn}) and delete training images?"):
        return False
    deleted = 0
    for fname in os.listdir(TRAINING_IMAGE_DIR):
        if fname.startswith(f"User.{label}.") or (f".{usn}." in fname):
            try:
                os.remove(os.path.join(TRAINING_IMAGE_DIR, fname))
                deleted += 1
            except Exception:
                pass
    df_new = df[~mask].copy()
    tmp = STUDENT_CSV + ".tmp"
    df_new.to_csv(tmp, index=False)
    try:
        os.replace(tmp, STUDENT_CSV)
    except Exception:
        df_new.to_csv(STUDENT_CSV, index=False)
    # Remove cached ORB index
    try:
        if os.path.exists(ORB_INDEX_PATH):
            os.remove(ORB_INDEX_PATH)
    except Exception:
        pass
    set_status(f"Removed {name} ({usn}). Deleted {deleted} training images.")
    messagebox.showinfo("Removed", f"Removed {name} ({usn}). Deleted {deleted} training images.")
    return True

# ------------------- Attendance loader -------------------
def load_all_attendance():
    if not os.path.exists(ATTENDANCE_DIR):
        return None
    frames = []
    for f in os.listdir(ATTENDANCE_DIR):
        if f.startswith("Attendance_") and f.endswith(".csv"):
            try:
                df = pd.read_csv(os.path.join(ATTENDANCE_DIR, f), dtype=str)
                frames.append(df)
            except Exception:
                continue
    if not frames:
        return None
    df_all = pd.concat(frames, ignore_index=True).fillna("")
    return df_all

# ------------------- DOCX retrieval (optional) -------------------
DEPT_SENTENCES = []
# (Docx loader omitted here for brevity; keep earlier code if you use it)

# ------------------- TTS & Voice -------------------
_tts_engine = None
PREFERRED_VOICE_NAMES_MALE = ["male", "david", "mark", "alex", "john", "daniel", "microsoft david", "microsoft zira"]

def _init_tts_engine_prefer_male():
    global _tts_engine
    try:
        _tts_engine = pyttsx3.init()
        voices = _tts_engine.getProperty("voices")
        chosen = None
        for v in voices:
            vname = (v.name or "").lower()
            vid = (v.id or "").lower()
            combined = vname + " " + vid
            for pref in PREFERRED_VOICE_NAMES_MALE:
                if pref in combined:
                    chosen = v
                    break
            if chosen:
                break
        if not chosen:
            for v in voices:
                vname = (v.name or "").lower()
                if "female" not in vname and "zira" not in vname:
                    chosen = v
                    break
        if chosen:
            try:
                _tts_engine.setProperty("voice", chosen.id)
            except Exception:
                pass
        _tts_engine.setProperty("rate", 170)
    except Exception as e:
        print("[TTS] init failed:", e)
        _tts_engine = None

_init_tts_engine_prefer_male()

def speak_text(text: str):
    if not text:
        return
    def _s(t):
        try:
            if _tts_engine:
                _tts_engine.say(t)
                _tts_engine.runAndWait()
        except Exception as e:
            print("[TTS] speak failed:", e)
    threading.Thread(target=_s, args=(text,), daemon=True).start()

# ------------------- Voice command processing -------------------
_sr_recognizer = sr.Recognizer()
_sr_microphone = None

def _normalize_text(s: str):
    return re.sub(r"[^\w\s]", "", s.lower()).strip()

def _get_all_registered_names():
    try:
        df = _load_student_df()
        return df["Name"].dropna().astype(str).tolist()
    except Exception:
        return []

def fuzzy_best_name_match(query, names, cutoff=0.6):
    query = query.lower().strip()
    best = None
    best_score = 0.0
    for n in names:
        score = difflib.SequenceMatcher(a=query, b=n.lower()).ratio()
        if score > best_score:
            best_score = score
            best = n
    if best_score >= cutoff:
        return best, best_score
    return None, best_score

def check_name_present_today_by_token(token: str):
    today = datetime.now().strftime("%Y-%m-%d")
    df_all = load_all_attendance()
    if df_all is None:
        return None
    df_today = df_all[df_all["Date"] == today]
    if df_today.empty:
        return None
    # exact contains match
    mask = df_today["Name"].fillna("").str.lower().str.contains(token.lower())
    if mask.any():
        return df_today[mask].to_dict("records")
    # check Id (USN) exact
    mask2 = df_today["Id"].fillna("").str.lower() == token.lower()
    if mask2.any():
        return df_today[mask2].to_dict("records")
    # fuzzy on registry names
    reg_names = _get_all_registered_names()
    best, score = fuzzy_best_name_match(token, reg_names, cutoff=0.45)
    if best:
        mask3 = df_today["Name"].fillna("").str.lower() == best.lower()
        if mask3.any():
            return df_today[mask3].to_dict("records")
    return None

def process_voice_command(text_norm: str, raw_text: str = None):
    raw_text = raw_text or text_norm
    q = text_norm.lower().strip()
    # document answers omitted here - keep existing logic if you need DEPT_SENTENCES
    m = re.search(r"\bis\s+([a-zA-Z\s]{2,60}?)\s+present\b", q)
    if m:
        name_token = m.group(1).strip()
        matches = check_name_present_today_by_token(name_token)
        if matches:
            first = matches[0]
            reply = f"Yes — {first.get('Name','')} is present, marked at {first.get('Time','')}."
            speak_text(reply); set_status(reply)
            return
        reg_names = _get_all_registered_names()
        best_name, score = fuzzy_best_name_match(name_token, reg_names, cutoff=0.45)
        if best_name:
            speak_text(f"Did you mean {best_name}? Checking...")
            set_status(f"Confirming: did you mean {best_name}?")
            # simple confirmation: try listening
            confirmed = _wait_for_confirmation(timeout=4)
            if not confirmed:
                speak_text("Please say the name again.")
                return
            matches2 = check_name_present_today_by_token(best_name)
            if matches2:
                first = matches2[0]
                reply = f"Yes — {first.get('Name','')} is present, marked at {first.get('Time','')}."
                speak_text(reply); set_status(reply); return
            else:
                reply = f"No attendance record found for {best_name} today."
                speak_text(reply); set_status(reply); return
        reply = f"No attendance record found for {name_token} today."
        speak_text(reply); set_status(reply); return

    if re.search(r"\bhow many\b.*\bpresent\b", q):
        df_all = load_all_attendance()
        if df_all is None:
            speak_text("No attendance data available."); set_status("No attendance data available."); return
        today = datetime.now().strftime("%Y-%m-%d")
        df_today = df_all[df_all["Date"] == today]
        n = int(df_today["Id"].nunique()) if not df_today.empty else 0
        reply = f"{n} students are present today."
        speak_text(reply); set_status(reply); return

    if re.search(r"\bwho(?:'s| is) present\b", q):
        df_all = load_all_attendance()
        if df_all is None:
            speak_text("No attendance data available."); set_status("No attendance data available."); return
        today = datetime.now().strftime("%Y-%m-%d")
        df_today = df_all[df_all["Date"] == today]
        if df_today.empty:
            speak_text("No one is marked present yet."); set_status("No one is marked present yet."); return
        names = df_today["Name"].unique().tolist()
        s = ", ".join(names[:12])
        if len(names) > 12:
            s += f", and {len(names)-12} more."
        reply = f"Present today: {s}"
        speak_text(reply); set_status(reply); return

    speak_text("Sorry, I can answer: 'Is <name> present?', 'How many present?', 'Who is present?'."); set_status("Voice fallback.")

# Simple confirmation listener (used by voice flow)
def _wait_for_confirmation(timeout=4):
    r = sr.Recognizer()
    try:
        with sr.Microphone() as src:
            r.adjust_for_ambient_noise(src, duration=0.3)
            audio = r.listen(src, timeout=timeout, phrase_time_limit=timeout)
        try:
            text = r.recognize_google(audio)
            text = _normalize_text(text)
            if any(w in text for w in ("yes", "yeah", "yup", "confirm", "correct", "sure")):
                return True
        except Exception:
            return False
    except Exception:
        return False
    return False

# ------------------- Voice loop & manual recording (unchanged concepts) -------------------
_voice_thread = None
_voice_stop_event = threading.Event()
WAKE_KEYWORDS = ["hey alan", "alan"]

def _voice_loop_dynamic():
    global _sr_microphone
    r = _sr_recognizer
    try:
        _sr_microphone = sr.Microphone()
    except Exception as e:
        set_status(f"Mic init failed: {e}")
        print("Mic init failed:", e); return
    with _sr_microphone as src:
        try:
            r.adjust_for_ambient_noise(src, duration=1.0)
        except Exception:
            pass
    set_status("Voice loop started. Say 'Hey Alan' to wake.")
    while not _voice_stop_event.is_set():
        try:
            with _sr_microphone as source:
                try:
                    audio = r.listen(source, timeout=1, phrase_time_limit=3)
                except sr.WaitTimeoutError:
                    continue
            try:
                text = r.recognize_google(audio)
            except sr.UnknownValueError:
                continue
            except sr.RequestError:
                set_status("Speech service unavailable."); time.sleep(1); continue
            if not text:
                continue
            text_norm = _normalize_text(text)
            clean_text = re.sub(r"[^\w\s]", " ", text_norm)
            woke = False
            for wk in WAKE_KEYWORDS:
                if re.search(r"\b" + re.escape(wk) + r"\b", clean_text):
                    woke = True; break
            if not woke and ("alan" in clean_text.split()):
                woke = True
            if not woke:
                continue
            # greeting
            greeting = "Hello — I am Alan. I am an attendance bot. How can I help?"
            try:
                if _tts_engine:
                    _tts_engine.say(greeting); _tts_engine.runAndWait()
                else:
                    speak_text(greeting)
            except Exception:
                speak_text(greeting)
            # wait for command
            with _sr_microphone as source:
                try:
                    r.adjust_for_ambient_noise(source, duration=0.4)
                except Exception:
                    pass
                r.pause_threshold = 1.0
                try:
                    audio_cmd = r.listen(source, timeout=20, phrase_time_limit=None)
                except sr.WaitTimeoutError:
                    speak_text("I didn't hear anything. Say Hey Alan to wake me."); set_status("No command heard"); continue
            try:
                cmd_text = r.recognize_google(audio_cmd)
            except sr.UnknownValueError:
                speak_text("Sorry, I couldn't understand that."); set_status("Unrecognized command."); continue
            except sr.RequestError:
                speak_text("Speech service is unavailable."); set_status("Speech recognition failed."); time.sleep(1); continue
            cmd_norm = _normalize_text(cmd_text)
            set_status(f"Voice command: {cmd_norm}")
            threading.Thread(target=process_voice_command, args=(cmd_norm, cmd_text), daemon=True).start()
        except Exception as e:
            print("Voice loop err:", e); time.sleep(0.5)
    set_status("Voice listener stopped.")

def start_voice_listener():
    global _voice_thread, _voice_stop_event
    if _voice_thread and _voice_thread.is_alive():
        return
    _voice_stop_event.clear()
    _voice_thread = threading.Thread(target=_voice_loop_dynamic, daemon=True)
    _voice_thread.start()
    set_status("Voice listener starting...")

def stop_voice_listener():
    global _voice_thread, _voice_stop_event
    _voice_stop_event.set()
    if _voice_thread:
        _voice_thread.join(timeout=1.0)
    _voice_thread = None
    set_status("Voice listener stopped.")

# Manual record (Start Query / Stop Query)
_manual_record_thread = None
_manual_record_stop_event = threading.Event()
_manual_record_frames = []
_manual_record_sample_rate = None
_manual_record_sample_width = None
_manual_record_lock = threading.Lock()

def _manual_recording_thread_func():
    global _manual_record_frames, _manual_record_sample_rate, _manual_record_sample_width
    r = sr.Recognizer()
    try:
        mic = sr.Microphone()
    except Exception as e:
        set_status(f"Manual recorder mic failed: {e}"); return
    with mic as source:
        try:
            r.adjust_for_ambient_noise(source, duration=0.6)
        except Exception:
            pass
    set_status("Manual recording started. Speak and press Stop when done.")
    with _manual_record_lock:
        _manual_record_frames = []
        _manual_record_sample_rate = None
        _manual_record_sample_width = None
    while not _manual_record_stop_event.is_set():
        try:
            with mic as source:
                audio = r.listen(source, timeout=2, phrase_time_limit=6)
            with _manual_record_lock:
                if _manual_record_sample_rate is None:
                    _manual_record_sample_rate = audio.sample_rate
                    _manual_record_sample_width = audio.sample_width
                if audio.sample_rate == _manual_record_sample_rate and audio.sample_width == _manual_record_sample_width:
                    _manual_record_frames.append(audio.get_raw_data())
        except sr.WaitTimeoutError:
            continue
        except Exception as e:
            print("Manual recorder err:", e); time.sleep(0.2); continue
    set_status("Manual recording stopped. Processing...")

def start_manual_recording():
    global _manual_record_thread, _manual_record_stop_event
    if _manual_record_thread and _manual_record_thread.is_alive():
        messagebox.showinfo("Recording", "Already recording.")
        return
    _manual_record_stop_event.clear()
    _manual_record_thread = threading.Thread(target=_manual_recording_thread_func, daemon=True)
    _manual_record_thread.start()
    set_status("Manual recording started...")

def stop_manual_recording_and_process():
    global _manual_record_thread, _manual_record_stop_event, _manual_record_frames, _manual_record_sample_rate, _manual_record_sample_width
    if not (_manual_record_thread and _manual_record_thread.is_alive()):
        messagebox.showinfo("Recording", "No active recording.")
        return
    _manual_record_stop_event.set()
    _manual_record_thread.join(timeout=4.0)
    _manual_record_thread = None
    with _manual_record_lock:
        frames = list(_manual_record_frames)
        sr_rate = _manual_record_sample_rate
        sr_width = _manual_record_sample_width
        _manual_record_frames = []
        _manual_record_sample_rate = None
        _manual_record_sample_width = None
    if not frames or sr_rate is None or sr_width is None:
        set_status("No audio recorded."); speak_text("I didn't record any audio."); return
    try:
        combined = b"".join(frames)
        audio_data = sr.AudioData(combined, sr_rate, sr_width)
    except Exception as e:
        print("Stitch audio failed:", e); set_status("Process failed"); speak_text("Failed to process recording."); return
    r = sr.Recognizer()
    try:
        set_status("Recognizing...")
        text = r.recognize_google(audio_data)
        set_status(f"You said: {text}")
        norm = _normalize_text(text)
        threading.Thread(target=process_voice_command, args=(norm, text), daemon=True).start()
    except sr.UnknownValueError:
        set_status("Could not understand."); speak_text("Sorry, I couldn't understand."); return
    except sr.RequestError as e:
        set_status("Speech service unavailable."); print("Manual SR error:", e); speak_text("Speech service unavailable."); return

# ------------------- ORB Indexing & Matching -------------------
def build_orb_index(save_cache=True):
    """
    Build ORB descriptors per label from all images in TRAINING_IMAGE_DIR.
    Format expectations: filenames like User.<label>.<USN>.<count>.jpg
    Returns:
      label_to_usn: {label: usn}
      label_to_name: {label: name}
      label_to_descs: {label: np.array([... descriptors ...])}  # stacked descriptors per label
    It also caches the index to ORB_INDEX_PATH (npz) for faster load on next run.
    """
    orb = cv2.ORB_create(nfeatures=ORB_N_FEATURES)
    df = _load_student_df()

    label_to_usn = {}
    label_to_name = {}
    for _, r in df.iterrows():
        try:
            lbl = int(r["Label"])
            label_to_usn[lbl] = r["USN"]
            label_to_name[lbl] = r["Name"]
        except Exception:
            continue

    # collect descriptors per label
    temp_desc = {}  # label -> list of descriptors arrays
    image_files = [p for p in glob.glob(os.path.join(TRAINING_IMAGE_DIR, "*")) if p.lower().endswith((".jpg",".jpeg",".png"))]
    for path in image_files:
        fname = os.path.basename(path)
        parts = fname.split(".")
        if len(parts) < 4:
            continue
        try:
            lbl = int(parts[1])
        except Exception:
            continue
        try:
            img = Image.open(path).convert("L")
            arr = np.array(img, dtype="uint8")
            arr = cv2.resize(arr, FACE_SIZE)
            kp, des = orb.detectAndCompute(arr, None)
            if des is None or len(kp) == 0:
                continue
            temp_desc.setdefault(lbl, []).append(des)
        except Exception:
            continue

    # stack descriptors per label to produce a single array per label for faster matching
    label_to_descs = {}
    for lbl, des_list in temp_desc.items():
        try:
            # vstack may fail for mismatched shapes; most descriptor arrays have same width (32)
            label_to_descs[lbl] = np.vstack(des_list)
        except Exception:
            # if stacking fails, keep as list of arrays (matching code will handle)
            label_to_descs[lbl] = des_list

    # optionally save cache
    if save_cache:
        try:
            # convert arrays to object arrays for np.savez
            np.savez_compressed(ORB_INDEX_PATH, label_to_usn=json.dumps(label_to_usn), label_to_name=json.dumps(label_to_name),
                                keys=np.array(list(label_to_descs.keys(),)), descs={str(k): v for k, v in label_to_descs.items()})
            # NOTE: np.savez doesn't like arbitrary dicts of arrays with non-uniform shapes in a straightforward way,
            # so as a robust fallback we'll write a small JSON metadata + per-label .npy files.
        except Exception:
            # robust fallback: write metadata and per-label .npy files
            meta = {"label_to_usn": label_to_usn, "label_to_name": label_to_name, "labels": []}
            try:
                os.makedirs(MODEL_DIR, exist_ok=True)
            except Exception:
                pass
            for lbl, arr in label_to_descs.items():
                meta["labels"].append(int(lbl))
                try:
                    np.save(os.path.join(MODEL_DIR, f"orb_desc_{lbl}.npy"), arr)
                except Exception:
                    pass
            try:
                with open(os.path.join(MODEL_DIR, "orb_index_meta.json"), "w", encoding="utf-8") as f:
                    json.dump(meta, f)
            except Exception:
                pass

    return label_to_usn, label_to_name, label_to_descs

def load_orb_index():
    """
    Try to load cached ORB index. If not present, build it.
    Returns the same triples as build_orb_index.
    """
    # Prefer simple meta + per-label file approach
    meta_path = os.path.join(MODEL_DIR, "orb_index_meta.json")
    if os.path.exists(meta_path):
        try:
            with open(meta_path, "r", encoding="utf-8") as f:
                meta = json.load(f)
            label_to_usn = meta.get("label_to_usn", {})
            label_to_name = meta.get("label_to_name", {})
            label_to_descs = {}
            for lbl in meta.get("labels", []):
                try:
                    arr = np.load(os.path.join(MODEL_DIR, f"orb_desc_{lbl}.npy"), allow_pickle=True)
                    label_to_descs[int(lbl)] = arr
                except Exception:
                    continue
            if label_to_descs:
                return label_to_usn, label_to_name, label_to_descs
        except Exception:
            pass

    # Try single-file cache (less reliable; kept for historic reasons)
    if os.path.exists(ORB_INDEX_PATH):
        try:
            packed = np.load(ORB_INDEX_PATH, allow_pickle=True)
            label_to_usn = json.loads(packed["label_to_usn"].tolist())
            label_to_name = json.loads(packed["label_to_name"].tolist())
            label_to_descs = {}
            # keys are stored as 'keys' and descs as separate arrays; this fragile branch may fail on some shapes
            try:
                keys = list(packed["keys"].tolist())
                for k in keys:
                    sk = str(k)
                    if sk in packed:
                        label_to_descs[int(k)] = packed[sk]
                if label_to_descs:
                    return label_to_usn, label_to_name, label_to_descs
            except Exception:
                pass
        except Exception:
            pass

    # fallback - rebuild index
    return build_orb_index(save_cache=True)

def _match_orb(des_query, des_db):
    """
    Match query descriptors to a database descriptor array.
    Uses BFMatcher with Hamming distance + crossCheck to get reliable matches.
    Returns number of "good" matches and list of matches.
    """
    if des_query is None or des_db is None:
        return 0, []
    # Use simple BFMatcher crossCheck for speed and determinism
    bf = cv2.BFMatcher(cv2.NORM_HAMMING, crossCheck=True)
    try:
        matches = bf.match(des_query, des_db)
    except Exception:
        return 0, []
    if not matches:
        return 0, []
    # good matches: those with distance <= ORB_GOOD_DISTANCE
    good = [m for m in matches if m.distance <= ORB_GOOD_DISTANCE]
    return len(good), good

# ------------------- Attendance register writer -------------------
def _register_filename_for(semester: str, subject: str):
    safe_sem = semester.replace(" ", "_").replace("/", "_")
    safe_subj = subject.replace(" ", "_").replace("/", "_")
    fname = f"Register_{safe_sem}_{safe_subj}.xlsx"
    return os.path.join(ATTENDANCE_REGISTERS_DIR, fname)

def seed_register_from_student_list(workbook, ws, semester):
    # attempt to fill USN/Name from StudentDetails.csv
    df = _load_student_df()
    rows = []
    if not df.empty:
        try:
            df_sem = df[df["Semester"].astype(str).str.lower() == semester.lower()]
            for _, r in df_sem.iterrows():
                rows.append((r["USN"], r["Name"]))
        except Exception:
            rows = []
    if not rows:
        return
    r = 2
    for usn, name in rows:
        ws.cell(row=r, column=1, value=str(usn))
        ws.cell(row=r, column=2, value=str(name))
        r += 1
    try:
        ws.column_dimensions[get_column_letter(1)].width = 20
        ws.column_dimensions[get_column_letter(2)].width = 30
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=1, column=2).alignment = Alignment(horizontal="center", vertical="center")
    except Exception:
        pass

def save_attendance_to_register(df_session: pd.DataFrame, semester: str, subject: str, class_name: str = ""):
    if df_session is None or df_session.empty:
        return False
    if Workbook is None or load_workbook is None:
        print("openpyxl not available"); return False
    try:
        df_session = df_session.copy()
        df_session["Id"] = df_session["Id"].astype(str)
        df_session["Name"] = df_session["Name"].astype(str)
    except Exception:
        pass

    today = datetime.now().strftime("%Y-%m-%d")
    file_path = _register_filename_for(semester, subject)
    created_new = False
    if not os.path.exists(file_path):
        wb = Workbook(); ws = wb.active; ws.title = "Attendance"
        ws.cell(row=1, column=1, value="USN"); ws.cell(row=1, column=2, value="Name")
        try:
            seed_register_from_student_list(wb, ws, semester)
        except Exception:
            pass
        created_new = True
    else:
        try:
            wb = load_workbook(file_path); ws = wb.active
        except Exception:
            wb = Workbook(); ws = wb.active; ws.title = "Attendance"
            ws.cell(row=1, column=1, value="USN"); ws.cell(row=1, column=2, value="Name")
            try:
                seed_register_from_student_list(wb, ws, semester)
            except Exception:
                pass
            created_new = True

    # find today's column
    max_col = ws.max_column
    date_col = None
    for c in range(3, max_col + 1):
        val = ws.cell(row=1, column=c).value
        if val and str(val).strip() == today:
            date_col = c; break
    if date_col is None:
        date_col = max_col + 1 if max_col >= 2 else 3
        ws.cell(row=1, column=date_col, value=today)
        try:
            ws.column_dimensions[get_column_letter(date_col)].width = 12
        except Exception:
            pass

    present_usns = set(df_session["Id"].astype(str).tolist())
    if ws.max_row < 2:
        r = 2
        for idx, nm in zip(df_session["Id"].astype(str), df_session["Name"].astype(str)):
            ws.cell(row=r, column=1, value=idx); ws.cell(row=r, column=2, value=nm); r += 1

    usn_to_row = {}
    for row in range(2, ws.max_row + 1):
        usn_val = ws.cell(row=row, column=1).value
        if usn_val is None:
            continue
        usn_to_row[str(usn_val).strip()] = row

    appended = 0
    for usn in present_usns:
        if usn not in usn_to_row:
            nm = df_session[df_session["Id"].astype(str) == usn]["Name"].iloc[0] if usn in df_session["Id"].astype(str).tolist() else ""
            new_row = ws.max_row + 1
            ws.cell(row=new_row, column=1, value=usn); ws.cell(row=new_row, column=2, value=nm)
            usn_to_row[usn] = new_row; appended += 1

    for usn, rownum in usn_to_row.items():
        cell = ws.cell(row=rownum, column=date_col)
        if usn in present_usns:
            cell.value = "P"
        else:
            cell.value = "A"

    tmp_path = file_path + ".tmp"
    try:
        wb.save(tmp_path); os.replace(tmp_path, file_path)
    except Exception:
        try:
            wb.save(file_path)
        except Exception as e:
            print("Save register failed:", e); return False

    set_status(f"Saved attendance into register: {os.path.basename(file_path)} (added {appended} new).")
    return True

# ------------------- Image capture & labeling -------------------
def take_images():
    """
    Capture training images and save as: User.<label>.<USN>.<count>.jpg
    This function behaves like your original capture routine.
    """
    global entry_usn, entry_name, entry_sem
    if entry_usn is None or entry_name is None:
        messagebox.showerror("Error", "Student fields not ready.")
        return
    usn = entry_usn.get().strip()
    name = entry_name.get().strip()
    sem = entry_sem.get().strip() if entry_sem else ""
    if not usn or not name:
        messagebox.showerror("Error", "Enter USN and Name.")
        return
    # ensure registered
    df = _load_student_df()
    mask = df["USN"].astype(str).str.lower() == usn.lower()
    if not mask.any():
        if not save_student_details(usn, name, sem):
            return
        df = _load_student_df()
        mask = df["USN"].astype(str).str.lower() == usn.lower()
    row = df[mask].iloc[0]
    label = str(row["Label"])
    if not os.path.exists(CASCADE_PATH):
        messagebox.showerror("Error", "haarcascade_frontalface_default.xml not found.")
        return
    cam = cv2.VideoCapture(0)
    if not cam.isOpened():
        messagebox.showerror("Error", "Cannot open camera.")
        return
    face_cascade = cv2.CascadeClassifier(CASCADE_PATH)
    count = 0
    set_status("Capturing images. Press Q to stop.")
    try:
        while True:
            ret, img = cam.read()
            if not ret:
                break
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(gray, 1.3, 5)
            # only save when exactly one face found to avoid mislabeling
            if len(faces) == 1:
                (x,y,w,h) = faces[0]
                face = gray[y:y+h, x:x+w]
                face = cv2.equalizeHist(face)
                try:
                    face = cv2.resize(face, FACE_SIZE)
                except Exception:
                    continue
                count += 1
                safe_usn = usn.replace(" ", "")
                fname = f"User.{label}.{safe_usn}.{count}.jpg"
                cv2.imwrite(os.path.join(TRAINING_IMAGE_DIR, fname), face)
                cv2.rectangle(img, (x,y),(x+w,y+h),(0,255,0),2)
            else:
                for (x,y,w,h) in faces:
                    cv2.rectangle(img, (x,y),(x+w,y+h),(0,0,255),2)
            cv2.imshow("Capturing - press Q to stop", img)
            if cv2.waitKey(1) & 0xFF == ord("q"):
                break
            if count >= 60:
                break
    finally:
        cam.release(); cv2.destroyAllWindows()
    set_status(f"Captured {count} images for {usn}.")
    messagebox.showinfo("Done", f"Captured {count} images for {usn}.")
    # invalidate ORB cache
    try:
        if os.path.exists(ORB_INDEX_PATH):
            os.remove(ORB_INDEX_PATH)
    except Exception:
        pass

# ------------------- Recognition: ORB-based attendance session -------------------
def track_images_orb():
    """
    ORB-based recognition routine. Builds/loads ORB index from training images and then
    recognizes faces in live camera, marking attendance accordingly.
    """
    global entry_sem, sem_var, subject_var, entry_class, current_faculty_name, current_role
    if entry_class is None or sem_var is None or subject_var is None:
        messagebox.showerror("Error", "Faculty fields not ready.")
        return
    class_name = entry_class.get().strip()
    semester = sem_var.get().strip()
    subject = subject_var.get().strip()
    if not semester:
        messagebox.showerror("Error", "Select Semester."); return
    if not subject or subject.startswith("("):
        messagebox.showerror("Error", "Select Subject."); return

    if not os.path.exists(CASCADE_PATH):
        messagebox.showerror("Error", "haarcascade_frontalface_default.xml not found.")
        return

    # Load or build ORB index
    label_to_usn, label_to_name, label_to_descs = load_orb_index()
    if not label_to_descs:
        messagebox.showerror("Error", "No training descriptors found. Capture faces first.")
        return

    face_cascade = cv2.CascadeClassifier(CASCADE_PATH)
    cam = cv2.VideoCapture(0)
    if not cam.isOpened():
        messagebox.showerror("Error", "Cannot open camera."); return

    # create ORB detector used at runtime to extract descriptors
    orb = cv2.ORB_create(nfeatures=ORB_N_FEATURES)

    session_start = datetime.now()
    session_start_str = session_start.strftime("%H:%M:%S")
    today = session_start.strftime("%Y-%m-%d")

    attendance = []
    marked_usns = set()
    last_seen = {}
    MIN_SECONDS_BETWEEN = 25

    set_status("Recognizing (ORB). Press Q in window to stop.")
    try:
        while True:
            ret, img = cam.read()
            if not ret:
                break
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(gray, 1.2, 5)
            for (x,y,w,h) in faces:
                face = gray[y:y+h, x:x+w]
                try:
                    face_resized = cv2.resize(face, FACE_SIZE)
                except Exception:
                    continue
                kp, des = orb.detectAndCompute(face_resized, None)
                recognized = False
                label_text = "Unknown"
                color = (0,0,255)
                if des is not None and len(kp) > 0:
                    # iterate all labels in index and find the best match (by good match count)
                    best_label = None
                    best_count = 0
                    best_ratio = 0.0
                    for lbl, des_db in label_to_descs.items():
                        # des_db is typically a stacked (N,32) array; ensure proper shape
                        try:
                            des_db_arr = np.array(des_db)
                        except Exception:
                            continue
                        if des_db_arr is None or des_db_arr.size == 0:
                            continue
                        good_count, matches = _match_orb(des, des_db_arr)
                        ratio = good_count / (len(kp) + 1e-9)
                        # choose label with highest good_count (and tie-break on ratio)
                        if good_count > best_count or (good_count == best_count and ratio > best_ratio):
                            best_count = good_count; best_ratio = ratio; best_label = lbl
                    # Evaluate acceptance criteria
                    if best_label is not None and (best_count >= ORB_MATCH_THRESHOLD or best_ratio >= ORB_RATIO_MIN):
                        usn = label_to_usn.get(best_label, str(best_label))
                        name = label_to_name.get(best_label, "")
                        label_text = f"{name} ({usn}) m:{best_count}"
                        recognized = True
                        color = (0,255,0)
                else:
                    # no descriptors detected in query face
                    recognized = False

                # If recognized and not recently seen, add to attendance
                if recognized:
                    # try to use usn and name variables; if not available parse label_text
                    try:
                        identifier_usn = usn
                        identifier_name = name
                    except Exception:
                        try:
                            identifier_usn = label_text.split("(")[-1].split(")")[0]
                            identifier_name = label_text.split("(")[0].strip()
                        except Exception:
                            identifier_usn = label_text
                            identifier_name = label_text
                    now = datetime.now()
                    last_t = last_seen.get(identifier_usn)
                    if last_t is None or (now - last_t).total_seconds() > MIN_SECONDS_BETWEEN:
                        last_seen[identifier_usn] = now
                        if identifier_usn not in marked_usns:
                            marked_usns.add(identifier_usn)
                            attendance.append({"Id": identifier_usn, "Name": identifier_name, "Class": class_name, "Semester": semester, "Subject": subject, "Faculty": current_faculty_name or "", "Date": today, "Time": now.strftime("%H:%M:%S"), "SessionStart": session_start_str, "SessionEnd": ""})

                cv2.rectangle(img, (x,y), (x+w,y+h), color, 2)
                cv2.putText(img, label_text, (x, y-8), cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 1)

            cv2.imshow("Recognizing - Press Q when done", img)
            if cv2.waitKey(1) & 0xFF == ord("q"):
                break
    finally:
        cam.release(); cv2.destroyAllWindows()

    # Save attendance if any
    if attendance:
        session_end_str = datetime.now().strftime("%H:%M:%S")
        for r in attendance:
            r["SessionEnd"] = session_end_str
        df_session = pd.DataFrame(attendance).drop_duplicates(subset=["Id"], keep="first")
        safe_class = (class_name or "").replace(" ", "")
        safe_sem = semester.replace(" ", "").replace("Semester","Sem")
        safe_subj = subject.replace(" ", "")
        timestamp_str = session_start.strftime("%Y-%m-%d_%H-%M-%S")
        fname = f"Attendance_{safe_sem}_{safe_subj}_{safe_class}_{timestamp_str}.csv"
        path = os.path.join(ATTENDANCE_DIR, fname)
        df_session.to_csv(path, index=False)
        try:
            saved = save_attendance_to_register(df_session, semester, subject, class_name)
        except Exception:
            saved = False
        msg = f"Attendance saved.\nSession file: {fname}"
        if saved:
            msg += f"\nUpdated register: {os.path.basename(_register_filename_for(semester,subject))}"
        messagebox.showinfo("Saved", msg)
        set_status(f"Attendance saved -> {fname}")
    else:
        messagebox.showinfo("Info", "No attendance captured.")
        set_status("No attendance captured.")

# ------------------- UI / Tkinter building -------------------
root = tk.Tk()
root.title("AIML Department Attendance Robot - ORB (no cv2.face)")
root.geometry("1000x640")
root.configure(bg=BG_MAIN)
root.grid_rowconfigure(0, weight=1); root.grid_columnconfigure(0, weight=1)

status_var = tk.StringVar(); status_var.set("Ready.")
def set_status(msg: str):
    try:
        status_var.set(msg); root.update_idletasks()
    except Exception:
        print("Status:", msg)

# Header
header = tk.Frame(root, bg=BG_MAIN)
header.pack(fill="x", padx=16, pady=(12,6))
tk.Label(header, text="AIML DEPARTMENT ROBOT - Alan (ORB fallback)", font=TITLE_FONT, bg=BG_MAIN, fg=FG_MAIN).pack(anchor="w")

# Main content
main = tk.Frame(root, bg=BG_MAIN)
main.pack(fill="both", expand=True, padx=16, pady=6)

# Left card: controls
left = tk.Frame(main, bg=BG_CARD, bd=1, relief="solid")
left.pack(side="left", fill="y", padx=8, pady=8)
tk.Label(left, text="Voice Controls", font=LABEL_FONT, bg=BG_CARD, fg=FG_MAIN).pack(anchor="w", padx=8, pady=(8,4))

btn_start_listener = tk.Button(left, text="Start Listener", command=start_voice_listener, bg=BUTTON_PRIMARY, fg=BUTTON_FG, font=BUTTON_FONT)
btn_start_listener.pack(fill="x", padx=8, pady=4)
btn_stop_listener = tk.Button(left, text="Stop Listener", command=stop_voice_listener, bg=BUTTON_PRIMARY, fg=BUTTON_FG, font=BUTTON_FONT)
btn_stop_listener.pack(fill="x", padx=8, pady=4)

tk.Label(left, text="Manual Query (Start/Stop)", bg=BG_CARD, fg=FG_MAIN, font=LABEL_FONT).pack(anchor="w", padx=8, pady=(8,2))
btn_start_query = tk.Button(left, text="Start Query (Record)", command=start_manual_recording, bg=BUTTON_PRIMARY, fg=BUTTON_FG, font=BUTTON_FONT)
btn_start_query.pack(fill="x", padx=8, pady=4)
btn_stop_query = tk.Button(left, text="Stop Query (Process)", command=stop_manual_recording_and_process, bg=BUTTON_PRIMARY, fg=BUTTON_FG, font=BUTTON_FONT)
btn_stop_query.pack(fill="x", padx=8, pady=4)

tk.Label(left, text="Quick Text Query:", bg=BG_CARD, fg=FG_MAIN, font=LABEL_FONT).pack(anchor="w", padx=8, pady=(8,2))
query_entry = tk.Entry(left, bg=ENTRY_BG, fg=ENTRY_FG)
query_entry.pack(fill="x", padx=8, pady=4)
def on_text_query():
    q = query_entry.get().strip()
    if not q:
        messagebox.showinfo("Query", "Type a query first.")
        return
    threading.Thread(target=process_voice_command, args=(_normalize_text(q), q), daemon=True).start()
tk.Button(left, text="Ask (text)", command=on_text_query, bg=BUTTON_PRIMARY, fg=BUTTON_FG, font=BUTTON_FONT).pack(fill="x", padx=8, pady=(0,8))

# Right card: attendance actions
right = tk.Frame(main, bg=BG_CARD, bd=1, relief="solid")
right.pack(side="left", fill="both", expand=True, padx=8, pady=8)
tk.Label(right, text="Attendance Quick Actions", font=LABEL_FONT, bg=BG_CARD, fg=FG_MAIN).pack(anchor="w", padx=8, pady=(8,4))

# faculty/student controls (simplified)
controls = tk.Frame(right, bg=BG_CARD)
controls.pack(anchor="nw", padx=8, pady=6, fill="x")

tk.Label(controls, text="Semester:", bg=BG_CARD, fg=FG_MAIN).grid(row=0, column=0, sticky="w", padx=4, pady=4)
sem_var = tk.StringVar(value=SEMESTER_OPTIONS[0])
sem_menu = ttk.Combobox(controls, textvariable=sem_var, values=SEMESTER_OPTIONS, state="readonly", width=22)
sem_menu.grid(row=0, column=1, padx=4, pady=4)

tk.Label(controls, text="Subject:", bg=BG_CARD, fg=FG_MAIN).grid(row=1, column=0, sticky="w", padx=4, pady=4)
subject_var = tk.StringVar(value="Select Subject")
subject_menu = ttk.Combobox(controls, textvariable=subject_var, values=SUBJECTS_BY_SEM.get(SEMESTER_OPTIONS[0], []), state="readonly", width=22)
subject_menu.grid(row=1, column=1, padx=4, pady=4)

tk.Label(controls, text="Class / Section:", bg=BG_CARD, fg=FG_MAIN).grid(row=2, column=0, sticky="w", padx=4, pady=4)
entry_class = tk.Entry(controls, bg=ENTRY_BG, fg=ENTRY_FG, width=24)
entry_class.grid(row=2, column=1, padx=4, pady=4)

# Student register capture area
reg_frame = tk.LabelFrame(right, text="Register Student & Capture Images", bg=BG_CARD, fg=FG_MAIN)
reg_frame.pack(fill="x", padx=8, pady=8)
tk.Label(reg_frame, text="USN:", bg=BG_CARD, fg=FG_MAIN).grid(row=0, column=0, padx=6, pady=4, sticky="w")
entry_usn = tk.Entry(reg_frame, bg=ENTRY_BG, fg=ENTRY_FG); entry_usn.grid(row=0, column=1, padx=6, pady=4)
tk.Label(reg_frame, text="Name:", bg=BG_CARD, fg=FG_MAIN).grid(row=1, column=0, padx=6, pady=4, sticky="w")
entry_name = tk.Entry(reg_frame, bg=ENTRY_BG, fg=ENTRY_FG); entry_name.grid(row=1, column=1, padx=6, pady=4)
tk.Label(reg_frame, text="Semester:", bg=BG_CARD, fg=FG_MAIN).grid(row=2, column=0, padx=6, pady=4, sticky="w")
entry_sem = tk.Entry(reg_frame, bg=ENTRY_BG, fg=ENTRY_FG); entry_sem.grid(row=2, column=1, padx=6, pady=4)

tk.Button(reg_frame, text="Save & Capture", command=take_images, bg=BUTTON_PRIMARY, fg=BUTTON_FG, font=BUTTON_FONT).grid(row=3, column=0, columnspan=2, pady=6, padx=8, sticky="ew")
tk.Button(reg_frame, text="Remove Student (USN)", command=lambda: remove_student_by_usn(entry_usn.get().strip()), bg=BUTTON_PRIMARY, fg=BUTTON_FG, font=BUTTON_FONT).grid(row=4, column=0, columnspan=2, pady=4, padx=8, sticky="ew")

# Attendance actions
tk.Button(right, text="Start Attendance (ORB)", command=track_images_orb, bg=BUTTON_PRIMARY, fg=BUTTON_FG, font=BUTTON_FONT).pack(padx=8, pady=(6,4), anchor="w")
tk.Button(right, text="Import Register (xlsx)", command=lambda: import_register_file(), bg=BUTTON_PRIMARY, fg=BUTTON_FG, font=BUTTON_FONT).pack(padx=8, pady=4, anchor="w")
tk.Button(right, text="Show Today's Counts", command=lambda: show_today_counts(), bg=BUTTON_PRIMARY, fg=BUTTON_FG, font=BUTTON_FONT).pack(padx=8, pady=4, anchor="w")

# small util functions used in UI
def import_register_file():
    if not load_workbook:
        messagebox.showwarning("Unavailable", "openpyxl not installed; register loading unavailable.")
        return
    f = filedialog.askopenfilename(title="Select register (xlsx)", filetypes=[("Excel", "*.xlsx *.xls")])
    if not f:
        return
    try:
        fname = os.path.basename(f)
        dest = os.path.join(ATTENDANCE_REGISTERS_DIR, fname)
        import shutil
        shutil.copyfile(f, dest)
        set_status(f"Register {fname} imported.")
    except Exception as e:
        messagebox.showerror("Error", str(e))

def show_today_counts():
    df_all = load_all_attendance()
    today = datetime.now().strftime("%Y-%m-%d")
    if df_all is None:
        messagebox.showinfo("Today", "No attendance data available."); return
    df_today = df_all[df_all["Date"] == today]
    present_n = int(df_today["Id"].nunique()) if not df_today.empty else 0
    total = None
    try:
        regs = []
        for f in os.listdir(ATTENDANCE_REGISTERS_DIR):
            if f.lower().endswith(".xlsx") and f.startswith("Register_"):
                try:
                    wb = load_workbook(os.path.join(ATTENDANCE_REGISTERS_DIR, f))
                    ws = wb.active
                    cnt = 0
                    for r in ws.iter_rows(min_row=2, max_col=1):
                        if r and r[0].value:
                            cnt += 1
                    regs.append(cnt)
                except Exception:
                    continue
        if regs:
            total = max(regs)
    except Exception:
        total = None
    if total is None:
        messagebox.showinfo("Today", f"{present_n} present today. No register size to compute absent.")
    else:
        absent = max(0, total - present_n)
        messagebox.showinfo("Today", f"Out of {total} students, {present_n} present, {absent} absent.")

# Status bar
status_bar = tk.Label(root, textvariable=status_var, bg=BUTTON_PRIMARY, fg=BUTTON_FG, anchor="w", padx=10)
status_bar.pack(side="bottom", fill="x")

# Minimal import helper: keep function defined after used
def _register_filename_for(semester: str, subject: str):
    safe_sem = semester.replace(" ", "_").replace("/", "_")
    safe_subj = subject.replace(" ", "_").replace("/", "_")
    fname = f"Register_{safe_sem}_{safe_subj}.xlsx"
    return os.path.join(ATTENDANCE_REGISTERS_DIR, fname)

# Start voice listener at startup (optional)
root.after(600, lambda: speak_text("Hi — I am Alan. I am an attendance bot."))  # greeting
start_voice_listener()

# Graceful shutdown
def on_closing():
    try:
        stop_voice_listener()
    except Exception:
        pass
    try:
        if _tts_engine:
            _tts_engine.stop()
    except Exception:
        pass
    try:
        root.destroy()
    except Exception:
        os._exit(0)

root.protocol("WM_DELETE_WINDOW", on_closing)

# Run mainloop
if __name__ == "__main__":
    set_status("Ready. Use Start Listener or Start Query.")
    root.mainloop()

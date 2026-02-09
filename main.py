# main.py
"""
Face Recognition Based Attendance System (UPDATED)
- Adds Start/Stop manual voice-query buttons inside the Chatbot window.
- Improved microphone calibration & wake-word tolerance in dynamic listener.
- Startup spoken greeting when the GUI opens.
- Manual Start/Stop recorder stitches audio chunks and uses Google SR, then routes into existing
  process_voice_command to speak attendance answers.
Notes:
- Requires: opencv-contrib-python, pyttsx3, SpeechRecognition, openpyxl, python-docx (optional), scikit-learn/joblib (for ML predictions)
- Place haarcascade_frontalface_default.xml in same folder as this file.
"""
import os
import re
import threading
import time
from datetime import datetime, timedelta
import difflib
import json

import cv2
import numpy as np
import pandas as pd
from PIL import Image

import tkinter as tk
from tkinter import messagebox, ttk

# speech & tts
import speech_recognition as sr
import pyttsx3

# docx reading
try:
    from docx import Document
except Exception:
    Document = None

# excel writer
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment
except Exception:
    Workbook = None
    load_workbook = None
    get_column_letter = None
    Alignment = None

# matplotlib for charts inside the app
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

# ===================== PATHS & DIRECTORIES =====================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

CASCADE_PATH = os.path.join(BASE_DIR, "haarcascade_frontalface_default.xml")

TRAINING_IMAGE_DIR = os.path.join(BASE_DIR, "TrainingImage")
TRAINER_DIR = os.path.join(BASE_DIR, "TrainingImageLabel")
TRAINER_PATH = os.path.join(TRAINER_DIR, "Trainer.yml")

STUDENT_CSV = os.path.join(BASE_DIR, "StudentDetails.csv")  # columns: Label,USN,Name,Semester
ATTENDANCE_DIR = os.path.join(BASE_DIR, "Attendance")
REPORT_DIR = os.path.join(BASE_DIR, "Reports")
MODEL_DIR = os.path.join(BASE_DIR, "Models")

DOCX_PATH = os.path.join(BASE_DIR, "COMPLETE DEPARTMENTAL GUIDE.docx")  # user uploaded doc

ATTENDANCE_REGISTERS_DIR = os.path.join(BASE_DIR, "AttendanceRegisters")

os.makedirs(TRAINING_IMAGE_DIR, exist_ok=True)
os.makedirs(TRAINER_DIR, exist_ok=True)
os.makedirs(ATTENDANCE_DIR, exist_ok=True)
os.makedirs(REPORT_DIR, exist_ok=True)
os.makedirs(MODEL_DIR, exist_ok=True)
os.makedirs(ATTENDANCE_REGISTERS_DIR, exist_ok=True)

# LBPH / FR settings
FACE_SIZE = (200, 200)
LBPH_THRESHOLD = 70.0   # lower = stricter (smaller conf values are better for LBPH)
FRAMES_FOR_CONFIRM = 8  # not heavily used in this version

# ===================== UI THEME =====================
TITLE_FONT = ("Helvetica", 22, "bold")
SUBTITLE_FONT = ("Helvetica", 16, "bold")
LABEL_FONT = ("Helvetica", 12)
BUTTON_FONT = ("Helvetica", 12, "bold")
SMALL_FONT = ("Helvetica", 10)

BG_DARK = "#DCEAFF"      # Light blue
FG_MAIN = "#000000"      # Black

BUTTON_BG = "#ffffff"
BUTTON_ACTIVE_BG = "#e2e8f0"
BUTTON_FG = "#000000"

ENTRY_BG = "#ffffff"
ENTRY_FG = "#000000"

# ===================== SEMESTER / SUBJECT / FACULTY =====================

SEMESTER_OPTIONS = ["3rd Semester", "5th Semester", "7th Semester"]

SUBJECTS_BY_SEM = {
    "3rd Semester": [
        "DMGT", "DSC (IPCC)", "OOPS (IPCC)", "DSDO", "PAI", "SC & R", "DSDO LAB",
        "DSC LAB", "OOPS LAB", "AEC (DVP)", "NSS", "PHY", "POC (Theory)", "C & Lab",
    ],
    "5th Semester": [
        "DS", "PML", "RMIPR", "DBMS", "HCI", "AIML LAB", "DBMS LAB",
        "MINI PROJECT", "ES", "PHY/YOGA",
    ],
    "7th Semester": [
        "BDA", "GEN AI", "CN", "OE-II", "MAJ PROJECT", "BDA LAB", "NLP LAB", "NLP",
    ],
}

TIMETABLE = {
    "Monday": {
        "3rd Semester": ["DMGT", "DSC (IPCC)", "OOPS (IPCC)"],
        "5th Semester": ["DS", "DBMS", "HCI"],
        "7th Semester": ["BDA", "GEN AI", "CN"],
    },
    "Tuesday": {
        "3rd Semester": ["PAI", "SC & R", "DSDO"],
        "5th Semester": ["PML", "DBMS LAB", "AIML LAB"],
        "7th Semester": ["BDA LAB", "NLP LAB", "MAJ PROJECT"],
    },
}


def get_today_timetable_for_sem(semester: str):
    day = datetime.now().strftime("%A")
    sem_table = TIMETABLE.get(day, {})
    return sem_table.get(semester, [])


# ----------------- FACULTY_INFO & AUTH (passwords set to "123") -------------
# Please update display_name and sem_subjects/sections if needed
FACULTY_INFO = {
    "kaveri_faculty": {
        "display_name": "Prof. Kaveri K",
        "sem_subjects": {"3rd Semester": ["PAI", "SC & R"], "7th Semester": ["BDA", "BDA LAB", "NLP LAB"],},
        "sections": {"3rd Semester": "III", "7th Semester": "VII"},
    },
    "sowmya_faculty": {
        "display_name": "Prof. Sowmya G",
        "sem_subjects": {"3rd Semester": ["DMGT", "SC & R"],},
        "sections": {"3rd Semester": "III"},
    },
    "ashwini_faculty": {
        "display_name": "Prof. Ashwini H",
        "sem_subjects": {"3rd Semester": ["DSC (IPCC)"], "5th Semester": ["DBMS"]},
        "sections": {"3rd Semester": "III", "5th Semester": "V"},
    },
    "nikita_faculty": {
        "display_name": "Prof. Nikita S",
        "sem_subjects": {"3rd Semester": ["OOPS (IPCC)"], "5th Semester": ["DS"]},
        "sections": {"3rd Semester": "III", "5th Semester": "V"},
    },
    "rajshekhar_faculty": {
        "display_name": "Prof. Rajshekhar P",
        "sem_subjects": {"7th Semester": ["NLP LAB", "NLP"]},
        "sections": {"7th Semester": "VII"},
    },
    # Add or extend faculty entries as needed
}

AUTH = {
    "faculty": {
        # username: password + faculty_key
        "kaveri": {"password": "123", "faculty_key": "kaveri_faculty"},
        "sowmya": {"password": "123", "faculty_key": "sowmya_faculty"},
        "ashwini": {"password": "123", "faculty_key": "ashwini_faculty"},
        "nikita": {"password": "123", "faculty_key": "nikita_faculty"},
        "rajshekhar": {"password": "123", "faculty_key": "rajshekhar_faculty"},
        # add further usernames with password "123"
    },
    "hod": {"hod": {"password": "123", "name": "HOD-AIML"}},
    "principal": {"principal": {"password": "123", "name": "Principal"}},
}

# -------------------- Auth & context --------------------
current_faculty_name = None
current_faculty_sem_subjects = {}
current_faculty_sections = {}
current_role = None

# ===================== TK ROOT & STATUS =====================
root = tk.Tk()
root.title("Face Recognition Based Attendance System")
root.geometry("1100x650")
root.minsize(900, 550)

# attempt nicer ttk theme
style = ttk.Style(root)
try:
    style.theme_use("clam")
except Exception:
    pass

root.configure(bg=BG_DARK)
root.option_add("*Background", BG_DARK)
root.option_add("*Foreground", "black")

root.grid_rowconfigure(0, weight=1)
root.grid_columnconfigure(0, weight=1)

status_var = tk.StringVar()
status_var.set("Ready.")

status_bar = tk.Label(
    root,
    textvariable=status_var,
    bg=BG_DARK,
    fg="black",
    anchor="w",
    padx=12,
    font=SMALL_FONT,
)
status_bar.grid(row=1, column=0, sticky="ew")


def set_status(msg: str):
    status_var.set(msg)
    root.update_idletasks()


# ===================== GLOBAL UI HANDLES =====================
frames = {}
entry_usn = None
entry_name = None
entry_class = None  # section / class name

sem_var = None
subject_var = None
subject_menu = None

student_usn_entry = None  # for student frame

# ===================== CSV / DATA HELPERS =====================
def ensure_student_csv():
    """Ensure CSV exists with columns: Label,USN,Name,Semester"""
    if not os.path.exists(STUDENT_CSV):
        df = pd.DataFrame(columns=["Label", "USN", "Name", "Semester"])
        df.to_csv(STUDENT_CSV, index=False)
        print("[INFO] Created StudentDetails.csv")
        return

    try:
        df = pd.read_csv(STUDENT_CSV, dtype=str)
        if "Label" not in df.columns or "USN" not in df.columns or "Name" not in df.columns:
            raise ValueError("Bad columns")
    except Exception:
        df = pd.DataFrame(columns=["Label", "USN", "Name", "Semester"])
        df.to_csv(STUDENT_CSV, index=False)
        print("[INFO] Recreated StudentDetails.csv with correct columns")


ensure_student_csv()


def _load_student_df():
    ensure_student_csv()
    try:
        df = pd.read_csv(STUDENT_CSV, dtype=str)
    except Exception:
        df = pd.DataFrame(columns=["Label", "USN", "Name", "Semester"])
    return df


def save_student_details(usn: str, student_name: str, semester: str):
    """
    Save student details. USN is unique (string like 3PD22AI900).
    Assign an integer Label for face recognizer (auto-increment).
    """
    usn = usn.strip()
    student_name = student_name.strip()
    semester = semester.strip()
    if not usn or not student_name:
        messagebox.showerror("Error", "USN and Name are required.")
        return False

    # load
    df = _load_student_df()
    # if USN exists, don't add
    if ((df["USN"].astype(str).str.lower()) == usn.lower()).any():
        messagebox.showinfo("Info", f"USN {usn} is already registered.")
        return False

    # next numeric label
    if df.empty:
        next_label = 1
    else:
        try:
            labels = df["Label"].dropna().astype(int).tolist()
            next_label = max(labels) + 1 if labels else 1
        except Exception:
            next_label = 1

    new_row = {"Label": str(next_label), "USN": usn, "Name": student_name, "Semester": semester}
    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    tmp = STUDENT_CSV + ".tmp"
    df.to_csv(tmp, index=False)
    try:
        os.replace(tmp, STUDENT_CSV)
    except Exception:
        df.to_csv(STUDENT_CSV, index=False)
    set_status(f"Student {usn} - {student_name} saved (label {next_label}).")
    return True


def remove_student_by_usn(usn: str):
    """Remove a student from StudentDetails and delete training images. Also optionally delete trainer to force retrain."""
    usn = usn.strip()
    if not usn:
        messagebox.showerror("Error", "Please enter USN to remove.")
        return False

    df = _load_student_df()
    mask = df["USN"].astype(str).str.lower() == usn.lower()
    if not mask.any():
        messagebox.showinfo("Info", f"No student with USN {usn} found.")
        return False

    matched = df[mask].iloc[0]
    label = str(matched["Label"])
    name = matched["Name"]

    if not messagebox.askyesno("Confirm Delete", f"Are you sure you want to remove {name} ({usn}) from records and delete their training images?"):
        return False

    # delete training images for that label & usn
    deleted = 0
    for fname in os.listdir(TRAINING_IMAGE_DIR):
        # filenames are User.<label>.<USN>.<count>.jpg
        if fname.startswith(f"User.{label}.") or (f".{usn}." in fname):
            try:
                os.remove(os.path.join(TRAINING_IMAGE_DIR, fname))
                deleted += 1
            except Exception:
                pass

    # remove from student CSV
    df_new = df[~mask].copy()
    tmp = STUDENT_CSV + ".tmp"
    df_new.to_csv(tmp, index=False)
    try:
        os.replace(tmp, STUDENT_CSV)
    except Exception:
        df_new.to_csv(STUDENT_CSV, index=False)

    # remove trainer model to avoid stale labels (user must retrain)
    if os.path.exists(TRAINER_PATH):
        try:
            os.remove(TRAINER_PATH)
        except Exception:
            pass

    set_status(f"Removed {name} ({usn}). Deleted {deleted} training images. Trainer removed; retrain model.")
    messagebox.showinfo("Removed", f"Removed {name} ({usn}).\nDeleted {deleted} training images.\nTrainer removed; please retrain the model.")
    return True


def load_all_attendance():
    if not os.path.exists(ATTENDANCE_DIR):
        return None

    frames_list = []
    for f in os.listdir(ATTENDANCE_DIR):
        if f.startswith("Attendance_") and f.endswith(".csv"):
            try:
                df = pd.read_csv(os.path.join(ATTENDANCE_DIR, f), dtype=str)
                frames_list.append(df)
            except Exception:
                continue

    if not frames_list:
        return None

    df_all = pd.concat(frames_list, ignore_index=True)

    for col in ["Id", "Name", "Class", "Semester", "Subject", "Faculty", "Date", "Time", "SessionStart", "SessionEnd"]:
        if col not in df_all.columns:
            df_all[col] = ""

    df_all = df_all.fillna("")
    return df_all


# ===================== DOCX LOADER & SIMPLE RETRIEVER =====================
DEPT_TEXT = ""
DEPT_SENTENCES = []

def load_department_docx(path=DOCX_PATH):
    global DEPT_TEXT, DEPT_SENTENCES
    if not os.path.exists(path):
        DEPT_TEXT = ""
        DEPT_SENTENCES = []
        return False
    if Document is None:
        DEPT_TEXT = ""
        DEPT_SENTENCES = []
        return False
    try:
        doc = Document(path)
        paragraphs = [p.text.strip() for p in doc.paragraphs if p.text and p.text.strip()]
        text = "\n".join(paragraphs)
        DEPT_TEXT = text
        # Split into sentences for retrieval (simple split)
        sentences = re.split(r'(?<=[\.\?\!])\s+', DEPT_TEXT)
        sentences = [s.strip() for s in sentences if s.strip()]
        DEPT_SENTENCES = sentences
        return True
    except Exception as e:
        DEPT_TEXT = ""
        DEPT_SENTENCES = []
        print("Docx load failed:", e)
        return False

# attempt load at startup
_doc_loaded = load_department_docx()

def answer_from_doc(query: str, max_chars=800):
    """
    Simple retrieval: match query tokens against departmental sentences and return best matches.
    """
    if not DEPT_SENTENCES:
        return None
    q = re.sub(r"[^\w\s]", " ", query.lower()).strip()
    if not q:
        return None
    tokens = [t for t in q.split() if len(t) > 2]
    if not tokens:
        tokens = q.split()
    scored = []
    for s in DEPT_SENTENCES:
        s_low = s.lower()
        score = sum(1 for t in tokens if t in s_low)
        if score > 0:
            scored.append((score, s))
    if not scored:
        best = None
        best_score = 0.0
        for s in DEPT_SENTENCES:
            score = difflib.SequenceMatcher(a=q, b=s.lower()).ratio()
            if score > best_score:
                best_score = score
                best = s
        if best_score > 0.25:
            return best[:max_chars]
        return None
    scored.sort(key=lambda x: x[0], reverse=True)
    top = [s for _, s in scored[:3]]
    out = " ".join(top)
    return out[:max_chars]


# ===================== VOICE & TTS SETUP (male-preferred) =====================
_voice_state = {
    "active": False,
    "active_until": None,
    "last_reply_at": None,
    "last_text": None,
}
# keep it longer so assistant stays listening after wake
VOICE_ACTIVE_TIMEOUT = 20
WAKE_KEYWORDS = ["hey alan", "hey allen", "alan", "allen"]
MIN_WORDS_FOR_COMMAND = 1
VOICE_REPLY_COOLDOWN = 3
CONFIRMATION_TIMEOUT = 4

_voice_thread = None
_voice_stop_event = threading.Event()

_tts_engine = None

# prefer male voice ids/names (common substrings)
PREFERRED_VOICE_NAMES_MALE = ["male", "david", "mark", "alex", "john", "daniel", "microsoft david", "microsoft zira"]


def _init_tts_engine_prefer_male():
    global _tts_engine
    try:
        _tts_engine = pyttsx3.init()
        voices = _tts_engine.getProperty("voices")
        chosen = None
        for v in voices:
            vname = (v.name or "").lower()
            vid = (v.id or "").lower()
            combined = vname + " " + vid
            for pref in PREFERRED_VOICE_NAMES_MALE:
                if pref.lower() in combined:
                    chosen = v
                    break
            if chosen:
                break
        # fallback: prefer voices not detected as female (best-effort)
        if not chosen:
            for v in voices:
                vname = (v.name or "").lower()
                if "female" not in vname and "zira" not in vname:
                    chosen = v
                    break
        if chosen:
            try:
                _tts_engine.setProperty("voice", chosen.id)
            except Exception:
                pass
        _tts_engine.setProperty("rate", 170)
    except Exception as e:
        print("[TTS] pyttsx3 init failed:", e)
        _tts_engine = None

_init_tts_engine_prefer_male()


def speak_text(text: str):
    """Speak asynchronous (non-blocking) — used for normal replies."""
    if not text:
        return
    def _speak(t):
        try:
            if _tts_engine:
                _tts_engine.say(t)
                _tts_engine.runAndWait()
        except Exception as e:
            print("[TTS] speak failed:", e)
    threading.Thread(target=_speak, args=(text,), daemon=True).start()


def speak_and_wait(text: str, fallback_set_active=True):
    """
    SPEAK synchronously in the current thread and return when done.
    This is used for the greeting so the listening timeout starts AFTER greeting finishes.
    If the TTS engine is not available, and fallback_set_active=True, we still mark assistant active.
    """
    global _tts_engine
    if not text:
        return False
    if _tts_engine:
        try:
            # Directly call runAndWait in the current thread so it blocks until speech is finished.
            _tts_engine.say(text)
            _tts_engine.runAndWait()
            return True
        except Exception as e:
            print("[TTS] speak_and_wait failed:", e)
            if fallback_set_active:
                return False
            return False
    else:
        # no TTS available: nothing to wait for
        return False


# -------------------- fuzzy name helpers --------------------
def _get_all_registered_names():
    try:
        df = _load_student_df()
        names = df["Name"].dropna().astype(str).tolist()
        return names
    except Exception:
        return []


def fuzzy_best_name_match(query, names, cutoff=0.6):
    query = query.lower().strip()
    best = None
    best_score = 0.0
    for n in names:
        score = difflib.SequenceMatcher(a=query, b=n.lower()).ratio()
        if score > best_score:
            best_score = score
            best = n
    if best_score >= cutoff:
        return best, best_score
    return None, best_score


def check_name_present_today(name: str):
    today = datetime.now().strftime("%Y-%m-%d")
    matches = []

    if not os.path.exists(ATTENDANCE_DIR):
        return None

    df_rows = []
    for f in os.listdir(ATTENDANCE_DIR):
        if f.startswith("Attendance_") and f.endswith(".csv"):
            try:
                df = pd.read_csv(os.path.join(ATTENDANCE_DIR, f), dtype=str)
            except Exception:
                continue
            if "Date" not in df.columns:
                continue
            df_today = df[df["Date"] == today]
            if not df_today.empty:
                df_rows.append(df_today)
    if not df_rows:
        return None

    df_all = pd.concat(df_rows, ignore_index=True)

    df_match = df_all[df_all["Name"].str.lower().str.contains(name.lower(), na=False)]
    if not df_match.empty:
        for _, r in df_match.iterrows():
            matches.append({"Id": r.get("Id", ""), "Name": r.get("Name", ""), "Time": r.get("Time", ""), "Subject": r.get("Subject", ""), "Class": r.get("Class", "")})
        return matches

    reg_names = _get_all_registered_names()
    best_name, score = fuzzy_best_name_match(name, reg_names, cutoff=0.55)
    if best_name:
        df_match2 = df_all[df_all["Name"].str.lower() == best_name.lower()]
        if not df_match2.empty:
            for _, r in df_match2.iterrows():
                matches.append({"Id": r.get("Id", ""), "Name": r.get("Name", ""), "Time": r.get("Time", ""), "Subject": r.get("Subject", ""), "Class": r.get("Class", "")})
            return matches
    return None

# ===================== Speech recognition dynamic listening thread =====================
_sr_recognizer = sr.Recognizer()
_sr_microphone = None

def log_voice_event(raw, norm, intent, match, score, resp):
    log_path = os.path.join(BASE_DIR, "voice_log.csv")
    row = {
        "timestamp": datetime.now().isoformat(),
        "raw": raw,
        "norm": norm,
        "intent": intent,
        "matched_name": match if match is not None else "",
        "score": score if score is not None else "",
        "response": resp if resp is not None else "",
    }
    df = pd.DataFrame([row])
    if not os.path.exists(log_path):
        df.to_csv(log_path, index=False)
    else:
        df.to_csv(log_path, mode="a", header=False, index=False)


def _normalize_text(s: str):
    return re.sub(r"[^\w\s]", "", s.lower()).strip()


def _wait_for_confirmation(timeout=4):
    r = sr.Recognizer()
    try:
        with sr.Microphone() as src:
            r.adjust_for_ambient_noise(src, duration=0.3)
            audio = r.listen(src, timeout=timeout, phrase_time_limit=timeout)
        try:
            text = r.recognize_google(audio)
            text = _normalize_text(text)
            if any(w in text for w in ("yes", "yeah", "yup", "confirm", "correct", "sure")):
                return True
        except Exception:
            return False
    except Exception:
        return False
    return False


def process_voice_command(text_norm: str, raw_text: str = None):
    """
    This function handles both standard attendance queries and documentary retrieval from the Word file.
    """
    raw_text = raw_text or text_norm
    set_status(f"Processing voice command: {text_norm}")
    q = text_norm.lower().strip()

    # First, check if the query is likely about departmental info
    doc_answer = answer_from_doc(q) if DEPT_SENTENCES else None
    if doc_answer:
        # respond with doc answer
        reply = doc_answer
        speak_text(reply)
        set_status("Answered from department guide.")
        log_voice_event(raw_text, text_norm, "doc_answer", None, None, reply)
        return

    # Attendance-related patterns
    m = re.search(r"\bis\s+([a-zA-Z\s]{2,60}?)\s+present\b", q)
    if m:
        name_token = m.group(1).strip()
        matches = check_name_present_today(name_token)
        if matches:
            first = matches[0]
            reply = f"Yes — {first['Name']} is present, marked at {first['Time']}."
            speak_text(reply); set_status(reply)
            log_voice_event(raw_text, text_norm, "is_present_exact", first['Name'], None, reply)
            return

        reg_names = _get_all_registered_names()
        best_name, score = fuzzy_best_name_match(name_token, reg_names, cutoff=0.45)
        if best_name:
            if score < 0.80:
                speak_text(f"Did you mean {best_name}? Please say yes or no.")
                set_status(f"Confirming: did you mean {best_name}?")
                confirmed = _wait_for_confirmation(timeout=CONFIRMATION_TIMEOUT)
                if not confirmed:
                    resp = "Please say the name again."
                    speak_text(resp)
                    log_voice_event(raw_text, text_norm, "confirm_needed", best_name, score, resp)
                    return
            matches2 = check_name_present_today(best_name)
            if matches2:
                first = matches2[0]
                reply = f"Yes — {first['Name']} is present, marked at {first['Time']}."
                speak_text(reply); set_status(reply)
                log_voice_event(raw_text, text_norm, "is_present_fuzzy_match", best_name, score, reply)
                return
            else:
                reply = f"No attendance record found for {best_name} today."
                speak_text(reply); set_status(reply)
                log_voice_event(raw_text, text_norm, "is_present_fuzzy_no_record", best_name, score, reply)
                return

        reply = f"No attendance record found for {name_token} today."
        speak_text(reply); set_status(reply)
        log_voice_event(raw_text, text_norm, "is_present_not_found", name_token, None, reply)
        return

    if re.search(r"\bhow many\b.*\bpresent\b", q):
        df_all = load_all_attendance()
        today = datetime.now().strftime("%Y-%m-%d")
        if df_all is None:
            reply = "No attendance data available."
            speak_text(reply); set_status(reply); log_voice_event(raw_text, text_norm, "how_many", None, None, reply); return
        df_today = df_all[df_all["Date"] == today]
        n = int(df_today["Id"].nunique())
        reply = f"{n} students are present today."
        speak_text(reply); set_status(reply); log_voice_event(raw_text, text_norm, "how_many", None, None, reply); return

    # New: how many absent? -> compute from register if register exists
    if re.search(r"\bhow many\b.*\babsent\b", q) or re.search(r"\bhow many\b.*\bare absent\b", q):
        # We attempt to compute based on current subject/semester register if present in memory from UI context.
        # Fallback to counting from known register files for today's date.
        df_all = load_all_attendance()
        today = datetime.now().strftime("%Y-%m-%d")
        if df_all is None:
            reply = "No attendance data available."
            speak_text(reply); set_status(reply); log_voice_event(raw_text, text_norm, "how_many_absent", None, None, reply); return
        df_today = df_all[df_all["Date"] == today]
        # If UI context exists (subject/semester), use it; otherwise try to infer class size from registers
        # Attempt to find a register file that matches subject/semester in the session data
        total_students = None
        # try: if there is any session today, use its Session file to find class size from register
        # fallback: if register file exists for any sem/subject, pick the largest
        # simpler approach: look for register files and find the largest number of USNs
        try:
            regs = []
            for f in os.listdir(ATTENDANCE_REGISTERS_DIR):
                if f.lower().endswith(".xlsx") and f.startswith("Register_"):
                    try:
                        wb = load_workbook(os.path.join(ATTENDANCE_REGISTERS_DIR, f))
                        ws = wb.active
                        # count rows with USN
                        cnt = 0
                        for r in ws.iter_rows(min_row=2, max_col=1):
                            if r and r[0].value:
                                cnt += 1
                        regs.append(cnt)
                    except Exception:
                        continue
            if regs:
                total_students = max(regs)
        except Exception:
            total_students = None

        present_n = int(df_today["Id"].nunique())
        if total_students is None:
            reply = f"{present_n} students are present today. I don't have register size to compute absent count."
            speak_text(reply); set_status(reply); log_voice_event(raw_text, text_norm, "how_many_absent_noreg", None, None, reply); return
        absent = max(0, total_students - present_n)
        reply = f"Out of {total_students} students, {present_n} are present today, so {absent} are absent."
        speak_text(reply); set_status(reply); log_voice_event(raw_text, text_norm, "how_many_absent", None, None, reply); return

    if re.search(r"\bwho(?:'s| is) present\b", q):
        df_all = load_all_attendance()
        today = datetime.now().strftime("%Y-%m-%d")
        if df_all is None:
            reply = "No attendance data available."
            speak_text(reply); set_status(reply); log_voice_event(raw_text, text_norm, "who_present", None, None, reply); return
        df_today = df_all[df_all["Date"] == today]
        names = df_today["Name"].unique().tolist()
        if not names:
            reply = "No one is marked present yet."
            speak_text(reply); set_status(reply); log_voice_event(raw_text, text_norm, "who_present_none", None, None, reply); return
        to_show = names[:12]
        s = ", ".join(to_show)
        if len(names) > 12:
            s += f", and {len(names)-12} more."
        reply = f"Present today: {s}"
        speak_text(reply); set_status(reply); log_voice_event(raw_text, text_norm, "who_present", None, None, reply); return

    # fallback
    reply = "Sorry, I can answer: 'Is <name> present?', 'How many present?', 'How many are absent?', 'Who is present?', or ask departmental questions (e.g., 'Who is HOD?', 'How many labs?')."
    speak_text(reply); set_status(reply); log_voice_event(raw_text, text_norm, "fallback", None, None, reply)
    return


def _voice_loop_dynamic():
    """
    Main voice thread:
    - Continuously listens for short ambient utterances; when a wake-word is detected,
      it greets synchronously, then *waits* for the user to actually speak (dynamic).
    - The dynamic listening uses recognizer.pause_threshold to decide end of phrase (silence).
    """
    global _sr_microphone, _voice_stop_event, _sr_recognizer
    r = _sr_recognizer
    try:
        _sr_microphone = sr.Microphone()
    except Exception as e:
        set_status(f"Microphone init failed: {e}")
        print("Mic init failed:", e)
        return

    # improved mic init & calibration
    r.dynamic_energy_threshold = False
    with _sr_microphone as source:
        try:
            # longer calibration helps in noisy rooms
            r.adjust_for_ambient_noise(source, duration=1.0)
            # set a safe baseline energy threshold
            r.energy_threshold = max(350, r.energy_threshold)
            r.pause_threshold = 0.8
        except Exception:
            pass

    set_status("Voice loop started (dynamic). Say 'Hey Alan' to wake.")
    while not _voice_stop_event.is_set():
        try:
            # short passive listen for ambient utterances (non-blocking-ish)
            with _sr_microphone as source:
                try:
                    audio = r.listen(source, timeout=1, phrase_time_limit=3)
                except sr.WaitTimeoutError:
                    continue
            try:
                text = r.recognize_google(audio)
            except sr.UnknownValueError:
                continue
            except sr.RequestError as e:
                set_status("Speech recognition service unavailable (network?).")
                print("RequestError from SR:", e)
                time.sleep(1)
                continue

            if not text:
                continue
            text_norm = _normalize_text(text)
            set_status(f"Voice heard: {text_norm}")

            # more tolerant wake detection: exact or contains short phrase
            woke = False
            clean_text = re.sub(r"[^\w\s]", " ", text_norm)
            for wk in WAKE_KEYWORDS:
                if re.search(r"\b" + re.escape(wk) + r"\b", clean_text):
                    woke = True
                    break
            # allow simpler "alan" anywhere
            if not woke and ("alan" in clean_text.split()):
                woke = True

            if not woke:
                continue

            # WAKE WORD detected
            set_status("Wake-word detected. Greeting...")
            greeting = "Hello — I am Alan. I am an attendance bot made by AIML students. How can I help you?"
            # speak greeting synchronously so the user hears it fully before we start recording
            # but do this in separate thread to avoid blocking the voice loop (we'll then listen)
            # We will use blocking here briefly to emulate previous behavior but not freeze GUI.
            try:
                speak_and_wait(greeting)
            except Exception:
                # fallback to async speak
                speak_text(greeting)

            # Now dynamic listening: wait for actual speech (no short phrase limit). We'll wait up to 20s for user to start,
            # then record until they pause (pause_threshold controls stop).
            with _sr_microphone as source:
                # recalibrate briefly for current ambient noise
                try:
                    r.adjust_for_ambient_noise(source, duration=0.4)
                except Exception:
                    pass
                # increase pause threshold so short pauses aren't cut off
                r.pause_threshold = 1.0   # seconds of silence that indicates end of phrase
                r.energy_threshold = max(300, r.energy_threshold)  # ensure threshold not too low; best-effort
                set_status("Listening for command (speak now)...")

                try:
                    # Wait up to 20s for user to start speaking; then record until silence (pause_threshold).
                    audio_cmd = r.listen(source, timeout=20, phrase_time_limit=None)
                except sr.WaitTimeoutError:
                    speak_text("I didn't hear anything. Say Hey Alan to wake me.")
                    set_status("No command heard after wake. Waiting again.")
                    continue

            # Recognize the collected command
            try:
                cmd_text = r.recognize_google(audio_cmd)
            except sr.UnknownValueError:
                speak_text("Sorry, I couldn't understand that. Please say it again.")
                set_status("Unrecognized command.")
                continue
            except sr.RequestError:
                speak_text("Speech service is unavailable right now.")
                set_status("Speech recognition request failed.")
                time.sleep(1)
                continue

            cmd_norm = _normalize_text(cmd_text)
            set_status(f"Command captured: {cmd_norm}")
            # Process command in background thread to keep loop responsive
            threading.Thread(target=process_voice_command, args=(cmd_norm, cmd_text), daemon=True).start()

        except Exception as e:
            print("Voice loop error:", e)
            time.sleep(0.5)

    set_status("Voice listener stopped.")


def start_voice_listener():
    global _voice_thread, _voice_stop_event
    if _voice_thread and _voice_thread.is_alive():
        return
    _voice_stop_event.clear()
    _voice_thread = threading.Thread(target=_voice_loop_dynamic, daemon=True)
    _voice_thread.start()
    set_status("Voice listener starting...")


def stop_voice_listener():
    global _voice_thread, _voice_stop_event
    _voice_stop_event.set()
    if _voice_thread:
        _voice_thread.join(timeout=1.0)
    _voice_thread = None
    set_status("Voice listener stopped.")


# ---------- manual Start/Stop recording globals ----------
_manual_record_thread = None
_manual_record_stop_event = threading.Event()
_manual_record_frames = []         # list of bytes frames
_manual_record_sample_rate = None
_manual_record_sample_width = None
_manual_record_lock = threading.Lock()

def _manual_recording_thread_func():
    """
    Background thread to capture multiple audio chunks until stop event is set.
    Each chunk is appended as bytes. After stop, the main thread will stitch them.
    """
    global _manual_record_stop_event, _manual_record_frames, _manual_record_sample_rate, _manual_record_sample_width
    r = sr.Recognizer()
    try:
        mic = sr.Microphone()
    except Exception as e:
        set_status(f"Manual recorder: microphone init failed: {e}")
        print("Manual recorder mic init failed:", e)
        return

    # calibrate briefly
    with mic as source:
        try:
            r.adjust_for_ambient_noise(source, duration=0.6)
            # set safe thresholds
            r.energy_threshold = max(300, r.energy_threshold)
            r.pause_threshold = 0.8
        except Exception:
            pass

    set_status("Manual recording started. Speak now and click Stop when done.")
    print("[Manual Recorder] started")

    # clear previous frames
    with _manual_record_lock:
        _manual_record_frames = []
        _manual_record_sample_rate = None
        _manual_record_sample_width = None

    while not _manual_record_stop_event.is_set():
        try:
            with mic as source:
                # listen for a short phrase chunk (non-blocking long)
                # phrase_time_limit keeps chunks manageable
                audio = r.listen(source, timeout=2, phrase_time_limit=6)
            # append raw bytes
            with _manual_record_lock:
                if _manual_record_sample_rate is None:
                    _manual_record_sample_rate = audio.sample_rate
                    _manual_record_sample_width = audio.sample_width
                # Only append if sample_rate matches initial chunk; otherwise drop incompatible chunk
                if audio.sample_rate == _manual_record_sample_rate and audio.sample_width == _manual_record_sample_width:
                    _manual_record_frames.append(audio.get_raw_data())
                else:
                    print("[Manual Recorder] skipped incompatible chunk (sample rate/width mismatch)")
        except sr.WaitTimeoutError:
            # nothing heard in that short window — keep looping until stop pressed
            continue
        except Exception as e:
            print("[Manual Recorder] error while listening:", e)
            time.sleep(0.2)
            continue

    set_status("Manual recording stopped. Processing...")
    print("[Manual Recorder] stopped")


def start_manual_recording():
    """Call when Start Query button clicked."""
    global _manual_record_thread, _manual_record_stop_event
    if _manual_record_thread and _manual_record_thread.is_alive():
        messagebox.showinfo("Recording", "Already recording. Click Stop to finish.")
        return
    _manual_record_stop_event.clear()
    _manual_record_thread = threading.Thread(target=_manual_recording_thread_func, daemon=True)
    _manual_record_thread.start()
    set_status("Manual recording started...")


def stop_manual_recording_and_process():
    """
    Call when Stop Query button clicked.
    This stops the recording thread, stitches audio chunks, runs speech recognition and processes the command.
    """
    global _manual_record_thread, _manual_record_stop_event, _manual_record_frames, _manual_record_sample_rate, _manual_record_sample_width

    if not (_manual_record_thread and _manual_record_thread.is_alive()):
        messagebox.showinfo("Recording", "No active recording session. Click Start Query first.")
        return

    # signal stop
    _manual_record_stop_event.set()
    # wait briefly for thread to finish
    _manual_record_thread.join(timeout=4.0)
    _manual_record_thread = None

    # build AudioData from frames
    with _manual_record_lock:
        frames = list(_manual_record_frames)
        sr_rate = _manual_record_sample_rate
        sr_width = _manual_record_sample_width
        # reset stored frames for next time
        _manual_record_frames = []
        _manual_record_sample_rate = None
        _manual_record_sample_width = None

    if not frames or sr_rate is None or sr_width is None:
        set_status("No audio recorded.")
        speak_text("I didn't record any audio. Please try again.")
        return

    try:
        combined = b"".join(frames)
        audio_data = sr.AudioData(combined, sr_rate, sr_width)
    except Exception as e:
        print("Failed to stitch audio chunks:", e)
        set_status("Failed to process recording.")
        speak_text("Failed to process the recording. Please try again.")
        return

    # Recognize
    r = sr.Recognizer()
    try:
        set_status("Recognizing your question...")
        text = r.recognize_google(audio_data)
        set_status(f"You said: {text}")
        print("[Manual Recognition] Transcribed:", text)
    except sr.UnknownValueError:
        set_status("Could not understand audio.")
        speak_text("Sorry, I couldn't understand that. Please try again.")
        return
    except sr.RequestError as e:
        set_status("Speech service unavailable (network?).")
        print("Manual recognizer RequestError:", e)
        speak_text("Speech service is currently unavailable.")
        return
    except Exception as e:
        print("Manual recognition error:", e)
        speak_text("An error occurred while recognizing audio.")
        return

    # Normalize and process the recognized text using existing command processor
    norm = _normalize_text(text)
    # process command in background thread
    threading.Thread(target=process_voice_command, args=(norm, text), daemon=True).start()


# ===================== Persistent Attendance Register (Excel) =====================
def _register_filename_for(semester: str, subject: str):
    """Return file path for the semester+subject register workbook."""
    safe_sem = semester.replace(" ", "_").replace("/", "_")
    safe_subj = subject.replace(" ", "_").replace("/", "_")
    fname = f"Register_{safe_sem}_{safe_subj}.xlsx"
    return os.path.join(ATTENDANCE_REGISTERS_DIR, fname)


def seed_register_from_student_list(workbook, ws, semester):
    """
    Optionally seed a new register sheet using:
      - StudentDetails.csv (registered students)
      - fallback: any uploaded semester excel list if present in BASE_DIR
    """
    # Primary: use student CSV
    df = _load_student_df()
    rows = []
    if not df.empty:
        try:
            df_sem = df[df["Semester"].astype(str).str.lower() == semester.lower()]
            for _, r in df_sem.iterrows():
                rows.append((r["USN"], r["Name"]))
        except Exception:
            rows = []

    # Secondary: if no students from CSV, try to find an uploaded semester-specific file in BASE_DIR
    if not rows:
        for f in os.listdir(BASE_DIR):
            if f.lower().endswith((".xlsx", ".xls")) and semester.split()[0].lower() in f.lower():
                try:
                    tmp = pd.read_excel(os.path.join(BASE_DIR, f), dtype=str)
                    if "USN" in tmp.columns and "Name" in tmp.columns:
                        tmp2 = tmp[["USN","Name"]].dropna()
                        rows = list(tmp2.itertuples(index=False, name=None))
                        break
                except Exception:
                    continue

    # If still empty, just leave workbook with header; user can add later
    if not rows:
        return

    # write USN/Name rows starting at row 2
    r = 2
    for usn, name in rows:
        ws.cell(row=r, column=1, value=str(usn))
        ws.cell(row=r, column=2, value=str(name))
        r += 1

    # set header alignment etc.
    ws.column_dimensions[get_column_letter(1)].width = 20
    ws.column_dimensions[get_column_letter(2)].width = 30
    for c in range(1, 3):
        ws.cell(row=1, column=c).alignment = Alignment(horizontal="center", vertical="center")


def save_attendance_to_register(df_session: pd.DataFrame, semester: str, subject: str, class_name: str = ""):
    """
    df_session: DataFrame with rows: Id (USN), Name, ... Time, Date
    semester & subject: strings
    - Creates/updates workbook: Register_<semester>_<subject>.xlsx
    - Adds today's date column (YYYY-MM-DD) if not present and fills P/A
    """
    if df_session is None or df_session.empty:
        return False

    if Workbook is None or load_workbook is None:
        # openpyxl not available
        print("openpyxl not available; cannot save register.")
        return False

    # ensure Id uses USN strings
    df_session = df_session.copy()
    df_session["Id"] = df_session["Id"].astype(str)
    df_session["Name"] = df_session["Name"].astype(str)

    today = datetime.now().strftime("%Y-%m-%d")
    file_path = _register_filename_for(semester, subject)
    created_new = False

    # if file doesn't exist, create and seed with student list
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        # header: USN, Name, then date columns
        ws.cell(row=1, column=1, value="USN")
        ws.cell(row=1, column=2, value="Name")
        # seed students
        try:
            seed_register_from_student_list(wb, ws, semester)
        except Exception:
            pass
        created_new = True
    else:
        try:
            wb = load_workbook(file_path)
            ws = wb.active
        except Exception:
            # if loading fails, create new
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance"
            ws.cell(row=1, column=1, value="USN")
            ws.cell(row=1, column=2, value="Name")
            try:
                seed_register_from_student_list(wb, ws, semester)
            except Exception:
                pass
            created_new = True

    # find today's column if exists
    max_col = ws.max_column
    date_col = None
    for c in range(3, max_col + 1):
        val = ws.cell(row=1, column=c).value
        if val and str(val).strip() == today:
            date_col = c
            break

    if date_col is None:
        # append new date as column
        date_col = max_col + 1 if max_col >= 2 else 3
        ws.cell(row=1, column=date_col, value=today)
        ws.column_dimensions[get_column_letter(date_col)].width = 12

    # build set of present USNs from df_session
    present_usns = set(df_session["Id"].astype(str).tolist())

    # determine total rows (USN list). If register had no students, attempt to seed from df_session
    max_row = ws.max_row
    if max_row < 2:
        # seed from session ids (use name mapping from session)
        r = 2
        for idx, nm in zip(df_session["Id"].astype(str), df_session["Name"].astype(str)):
            ws.cell(row=r, column=1, value=idx)
            ws.cell(row=r, column=2, value=nm)
            r += 1
        max_row = ws.max_row

    # Build a map of USN -> row number for faster writes
    usn_to_row = {}
    for row in range(2, ws.max_row + 1):
        usn_val = ws.cell(row=row, column=1).value
        if usn_val is None:
            continue
        usn_to_row[str(usn_val).strip()] = row

    # If any present ids not in register, append them at bottom
    appended = 0
    for usn in present_usns:
        if usn not in usn_to_row:
            # find name from df_session
            try:
                nm = df_session[df_session["Id"].astype(str) == usn]["Name"].iloc[0]
            except Exception:
                nm = ""
            new_row = ws.max_row + 1
            ws.cell(row=new_row, column=1, value=usn)
            ws.cell(row=new_row, column=2, value=nm)
            usn_to_row[usn] = new_row
            appended += 1

    # Now mark P/A for every USN present in the register
    for usn, rownum in usn_to_row.items():
        cell = ws.cell(row=rownum, column=date_col)
        if usn in present_usns:
            cell.value = "P"
        else:
            cell.value = "A"

    # save workbook atomically
    tmp_path = file_path + ".tmp"
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, file_path)
    except Exception as e:
        try:
            wb.save(file_path)
        except Exception as e2:
            print("Failed to save register workbook:", e2)
            return False

    set_status(f"Saved attendance into register: {os.path.basename(file_path)} (appended {appended} new students).")
    return True


# ===================== FACIAL DATA FUNCTIONS =====================
def take_images():
    """
    Capture images for a student. Inputs: USN (string), Name (string), Semester (string).
    Saves images as: User.<label>.<USN>.<count>.jpg
    """
    global entry_usn, entry_name, entry_class
    if entry_usn is None or entry_name is None:
        messagebox.showerror("Error", "Student entry fields are not ready.")
        return
    usn = entry_usn.get().strip()
    name = entry_name.get().strip()
    semester = entry_class.get().strip() if entry_class else ""

    if not usn or not name:
        messagebox.showerror("Error", "Please enter USN and Name.")
        return

    # If USN not registered, add
    df = _load_student_df()
    mask = df["USN"].astype(str).str.lower() == usn.lower()
    if not mask.any():
        added = save_student_details(usn, name, semester)
        if not added:
            return
        df = _load_student_df()
        mask = df["USN"].astype(str).str.lower() == usn.lower()

    row = df[mask].iloc[0]
    label = str(row["Label"])

    if not os.path.exists(CASCADE_PATH):
        messagebox.showerror("Error", "haarcascade_frontalface_default.xml not found.\nPlace it in the same folder as main.py.")
        return

    cam = cv2.VideoCapture(0)
    if not cam.isOpened():
        messagebox.showerror("Error", "Cannot open camera.")
        return
    face_cascade = cv2.CascadeClassifier(CASCADE_PATH)

    count = 0
    set_status("Capturing images. Press Q to stop earlier.")
    try:
        while True:
            ret, img = cam.read()
            if not ret:
                break
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(gray, 1.3, 5)
            # Only save when exactly one face is detected to avoid mislabeling
            if len(faces) == 1:
                (x, y, w, h) = faces[0]
                count += 1
                face = gray[y: y + h, x: x + w]
                face = cv2.equalizeHist(face)
                try:
                    face = cv2.resize(face, FACE_SIZE)
                except Exception:
                    continue
                safe_usn = usn.replace(" ", "")
                file_name = f"User.{label}.{safe_usn}.{count}.jpg"
                cv2.imwrite(os.path.join(TRAINING_IMAGE_DIR, file_name), face)
                cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)
            else:
                for (x, y, w, h) in faces:
                    cv2.rectangle(img, (x, y), (x + w, y + h), (0, 0, 255), 2)

            cv2.imshow("Capturing - Press Q to stop", img)
            if cv2.waitKey(1) & 0xFF == ord("q"):
                break
            if count >= 60:
                break
    finally:
        cam.release()
        cv2.destroyAllWindows()
    set_status(f"Captured {count} images for USN {usn}.")
    messagebox.showinfo("Done", f"Captured {count} images for student {usn}.")


def train_images():
    if not hasattr(cv2, "face") or not hasattr(cv2.face, "LBPHFaceRecognizer_create"):
        messagebox.showerror("OpenCV Error", "Your OpenCV build does not have LBPHFaceRecognizer.\nInstall opencv-contrib-python.")
        return

    image_paths = [os.path.join(TRAINING_IMAGE_DIR, f) for f in os.listdir(TRAINING_IMAGE_DIR) if f.lower().endswith((".jpg", ".jpeg", ".png"))]
    if not image_paths:
        messagebox.showerror("Error", "No training images found.")
        return

    face_samples = []
    labels = []
    for path in image_paths:
        try:
            pil_img = Image.open(path).convert("L")
        except Exception:
            continue
        img_np = np.array(pil_img, "uint8")
        fname = os.path.basename(path)
        parts = fname.split(".")
        # expected: User.<label>.<USN>.<count>.jpg
        if len(parts) < 4:
            continue
        try:
            lbl = int(parts[1])
        except Exception:
            continue
        try:
            face_resized = cv2.resize(img_np, FACE_SIZE)
        except Exception:
            continue
        face_samples.append(face_resized)
        labels.append(lbl)

    if not face_samples:
        messagebox.showerror("Error", "No valid training images.")
        return

    recognizer = cv2.face.LBPHFaceRecognizer_create()
    recognizer.train(face_samples, np.array(labels))

    # quick on-training evaluation
    correct = 0
    total = 0
    for sample, true_lbl in zip(face_samples, labels):
        try:
            pred_lbl, conf = recognizer.predict(sample)
        except Exception:
            pred_lbl, conf = -1, 999.0
        total += 1
        if pred_lbl == true_lbl:
            correct += 1
    acc = correct / total if total > 0 else 0.0

    # atomic save
    tmp_path = TRAINER_PATH + ".tmp"
    try:
        recognizer.save(tmp_path)
        os.replace(tmp_path, TRAINER_PATH)
    except Exception as e:
        messagebox.showerror("Save Error", f"Failed to save trainer: {e}")
        return

    set_status("Model trained successfully.")
    messagebox.showinfo("Success", f"Model trained successfully.\nTrain-set accuracy: {acc*100:.1f}%\nTraining samples: {len(labels)}")


def evaluate_trained_model(trainer_path=TRAINER_PATH, images_dir=TRAINING_IMAGE_DIR):
    """Evaluate trained model on labeled images and write diagnostics."""
    if not os.path.exists(trainer_path):
        messagebox.showerror("Evaluate", "No trained model found. Train the model first.")
        return

    if not getattr(cv2, "face", None) or not hasattr(cv2.face, "LBPHFaceRecognizer_create"):
        messagebox.showerror("Evaluate Error", "OpenCV face recognizer not available.")
        return

    recognizer = cv2.face.LBPHFaceRecognizer_create()
    try:
        recognizer.read(trainer_path)
    except Exception as e:
        messagebox.showerror("Evaluate Error", f"Failed to read trainer: {e}")
        return

    records = []
    image_files = [f for f in os.listdir(images_dir) if f.lower().endswith(('.jpg','.jpeg','.png'))]
    if not image_files:
        messagebox.showinfo("Evaluate", "No training images found to evaluate.")
        return

    id_list = []
    y_true = []
    y_pred = []
    for fname in image_files:
        path = os.path.join(images_dir, fname)
        parts = fname.split(".")
        if len(parts) < 4:
            continue
        try:
            true_lbl = int(parts[1])
            usn = parts[2]
        except Exception:
            continue
        try:
            img = Image.open(path).convert("L")
            arr = np.array(img, dtype="uint8")
            arr_resized = cv2.resize(arr, FACE_SIZE)
        except Exception:
            continue
        try:
            predicted_lbl, conf = recognizer.predict(arr_resized)
        except Exception:
            predicted_lbl, conf = -1, 999.0
        records.append({"file": fname, "true_lbl": true_lbl, "pred_lbl": predicted_lbl, "conf": conf, "usn": usn})
        y_true.append(true_lbl)
        y_pred.append(predicted_lbl)
        id_list.extend([true_lbl, predicted_lbl])

    df_eval = pd.DataFrame(records)
    eval_csv = os.path.join(BASE_DIR, "training_evaluation.csv")
    df_eval.to_csv(eval_csv, index=False)

    labels = sorted(list(set([x for x in id_list if isinstance(x, int) or str(x).isdigit()])))
    try:
        from sklearn.metrics import confusion_matrix
        cm = confusion_matrix(y_true, y_pred, labels=labels)
        cm_df = pd.DataFrame(cm, index=[f"true_{l}" for l in labels], columns=[f"pred_{l}" for l in labels])
        cm_csv = os.path.join(BASE_DIR, "training_confusion_matrix.csv")
        cm_df.to_csv(cm_csv)
    except Exception:
        pass

    confusions = df_eval[df_eval["true_lbl"] != df_eval["pred_lbl"]].groupby(["true_lbl", "pred_lbl"]).size().reset_index(name="count").sort_values("count", ascending=False)
    if confusions.empty:
        messagebox.showinfo("Evaluation", "No confusions detected on training images (perfect on-training accuracy).")
    else:
        top = confusions.head(20).to_dict(orient="records")
        msg_lines = ["Top confusions (true_label -> pred_label : count):"]
        for r in top:
            msg_lines.append(f"{r['true_lbl']} -> {r['pred_lbl']} : {r['count']}")
        messagebox.showinfo("Evaluation", "\n".join(msg_lines))
    set_status("Evaluation complete. See training_evaluation.csv and training_confusion_matrix.csv.")


def track_images():
    global entry_class, sem_var, subject_var
    if entry_class is None or sem_var is None or subject_var is None:
        messagebox.showerror("Error", "Faculty fields are not ready.")
        return
    class_name = entry_class.get().strip()
    semester = sem_var.get().strip()
    subject = subject_var.get().strip()
    if not semester:
        messagebox.showerror("Error", "Please select a Semester.")
        return
    if not subject or subject.startswith("("):
        messagebox.showerror("Error", "Please select a valid Subject.")
        return

    if current_faculty_sem_subjects and current_role == "faculty":
        allowed_subjects = current_faculty_sem_subjects.get(semester, [])
        if subject not in allowed_subjects:
            messagebox.showerror("Not Allowed", f"You are not assigned to teach {subject} in {semester}.\nPlease check your selection.")
            return

    if not os.path.exists(TRAINER_PATH):
        messagebox.showerror("Error", "Model not trained yet.\nPlease click 'Train Model' first.")
        return

    df_students = _load_student_df()
    if df_students.empty:
        messagebox.showerror("Error", "No student details found. Register students first.")
        return

    # build label -> (USN, Name) maps
    label_to_usn = {}
    label_to_name = {}
    for _, r in df_students.iterrows():
        try:
            lbl = int(r["Label"])
        except Exception:
            continue
        label_to_usn[lbl] = r["USN"]
        label_to_name[lbl] = r["Name"]

    if not hasattr(cv2, "face") or not hasattr(cv2.face, "LBPHFaceRecognizer_create"):
        messagebox.showerror("OpenCV Error", "Your OpenCV build does not have LBPHFaceRecognizer.")
        return

    recognizer = cv2.face.LBPHFaceRecognizer_create()
    try:
        recognizer.read(TRAINER_PATH)
    except Exception as e:
        messagebox.showerror("Model Error", "Trainer.yml is corrupted.\nDelete it, recapture faces and retrain.")
        print(e)
        return

    if not os.path.exists(CASCADE_PATH):
        messagebox.showerror("Error", "haarcascade_frontalface_default.xml not found.\nPlace it in the same folder as main.py.")
        return

    face_cascade = cv2.CascadeClassifier(CASCADE_PATH)

    cam = cv2.VideoCapture(0)
    if not cam.isOpened():
        messagebox.showerror("Error", "Cannot open camera.")
        return

    session_start_dt = datetime.now()
    session_start_str = session_start_dt.strftime("%H:%M:%S")

    attendance = []
    marked_usns = set()
    last_seen_time = {}

    today_str = session_start_dt.strftime("%Y-%m-%d")
    set_status("Recognizing... Press Q to stop when finished with the class.")

    MIN_SECONDS_BETWEEN_SAME_ID = 30

    try:
        while True:
            ret, img = cam.read()
            if not ret:
                break
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(gray, 1.2, 5)
            for (x, y, w, h) in faces:
                face_gray = gray[y:y + h, x:x + w]
                face_gray = cv2.equalizeHist(face_gray)
                try:
                    face_resized = cv2.resize(face_gray, FACE_SIZE)
                except Exception:
                    continue
                try:
                    label_pred, conf = recognizer.predict(face_resized)
                except Exception as e:
                    label_pred, conf = -1, 999.0

                if conf < LBPH_THRESHOLD and label_pred in label_to_usn:
                    usn = label_to_usn[label_pred]
                    name = label_to_name.get(label_pred, "")
                    color = (0, 255, 0)
                    label = f"{name} ({usn}) {conf:.1f}"
                    now = datetime.now()
                    last_time = last_seen_time.get(usn)
                    if last_time is None or (now - last_time).total_seconds() > MIN_SECONDS_BETWEEN_SAME_ID:
                        last_seen_time[usn] = now
                        if usn not in marked_usns:
                            marked_usns.add(usn)
                            attendance.append({"Id": usn, "Name": name, "Class": class_name, "Semester": semester, "Subject": subject, "Faculty": current_faculty_name or "", "Date": today_str, "Time": now.strftime("%H:%M:%S"), "SessionStart": session_start_str, "SessionEnd": ""})
                else:
                    color = (0, 0, 255)
                    label = f"Unknown ({conf:.1f})"
                cv2.rectangle(img, (x, y), (x + w, y + h), color, 2)
                cv2.putText(img, label, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)
            cv2.imshow("Recognizing - Press Q when done", img)
            if cv2.waitKey(1) & 0xFF == ord("q"):
                break
    finally:
        cam.release()
        cv2.destroyAllWindows()

    if attendance:
        session_end_str = datetime.now().strftime("%H:%M:%S")
        for row in attendance:
            row["SessionEnd"] = session_end_str

        df_session = pd.DataFrame(attendance)
        df_session = df_session.drop_duplicates(subset=["Id"], keep="first")

        safe_class = (class_name or "").replace(" ", "")
        safe_sem = semester.replace(" ", "").replace("Semester", "Sem")
        safe_subj = subject.replace(" ", "")

        timestamp_str = session_start_dt.strftime("%Y-%m-%d_%H-%M-%S")
        file_name = f"Attendance_{safe_sem}_{safe_subj}_{safe_class}_{timestamp_str}.csv"
        file_path = os.path.join(ATTENDANCE_DIR, file_name)
        df_session.to_csv(file_path, index=False)

        # Also save/update persistent register
        try:
            saved = save_attendance_to_register(df_session, semester, subject, class_name)
        except Exception as e:
            saved = False
            print("Register save failed:", e)

        message = f"Attendance saved.\nSession file: {file_name}"
        if saved:
            message += f"\nAlso updated register: {os.path.basename(_register_filename_for(semester,subject))}"
        messagebox.showinfo("Saved", message)
        set_status(f"Attendance saved -> {file_name}")
    else:
        messagebox.showinfo("Info", "No attendance captured.")
        set_status("No attendance captured.")


# ===================== STUDENT FUNCTIONS =====================
def get_usn_from_entry():
    global student_usn_entry
    if student_usn_entry is None:
        messagebox.showerror("Error", "Student USN field not ready.")
        return None
    usn = student_usn_entry.get().strip()
    if not usn:
        messagebox.showerror("Error", "Please enter your USN.")
        return None
    return usn


def student_check_today():
    usn = get_usn_from_entry()
    if usn is None:
        return
    today = datetime.now().strftime("%Y-%m-%d")
    records = []
    for f in os.listdir(ATTENDANCE_DIR):
        if f.startswith("Attendance_") and f.endswith(".csv"):
            try:
                df = pd.read_csv(os.path.join(ATTENDANCE_DIR, f), dtype=str)
            except Exception:
                continue
            if "Date" not in df.columns or "Id" not in df.columns:
                continue
            df_today = df[(df["Date"] == today) & (df["Id"].astype(str).str.lower() == usn.lower())]
            if not df_today.empty:
                records.append(df_today)
    if not records:
        messagebox.showinfo("Attendance", "No attendance found for today.")
        return
    df_all = pd.concat(records, ignore_index=True)
    lines = []
    for _, r in df_all.iterrows():
        lines.append(f"{r['Date']} {r['Time']} – {r.get('Semester','')} {r.get('Subject','')} ({r.get('Class','')})")
    messagebox.showinfo("Attendance", "Present today in:\n\n" + "\n".join(lines))


def student_view_history():
    usn = get_usn_from_entry()
    if usn is None:
        return
    records = []
    for f in os.listdir(ATTENDANCE_DIR):
        if f.startswith("Attendance_") and f.endswith(".csv"):
            try:
                df = pd.read_csv(os.path.join(ATTENDANCE_DIR, f), dtype=str)
            except Exception:
                continue
            if "Id" not in df.columns:
                continue
            df_sid = df[df["Id"].astype(str).str.lower() == usn.lower()]
            if not df_sid.empty:
                records.append(df_sid)
    if not records:
        messagebox.showinfo("Attendance", "No attendance records found.")
        return
    df_all = pd.concat(records, ignore_index=True)
    df_all = df_all.fillna("")
    df_all["Date_dt"] = pd.to_datetime(df_all["Date"], errors="coerce")
    df_all = df_all.sort_values(by=["Date_dt", "Time"], na_position="last")
    df_all.drop(columns=["Date_dt"], inplace=True)

    win = tk.Toplevel(root)
    win.title(f"Attendance History - {usn}")
    win.geometry("850x450")
    win.configure(bg=BG_DARK)
    win.grab_set()
    tk.Label(win, text=f"Attendance History – USN {usn}", font=SUBTITLE_FONT, bg=BG_DARK, fg=FG_MAIN).pack(pady=8)

    frame = tk.Frame(win, bg=BG_DARK)
    frame.pack(fill="both", expand=True, padx=10, pady=10)

    style = ttk.Style(win)
    style.theme_use("clam")
    style.configure("Student.Treeview", font=LABEL_FONT, rowheight=24, background="white", fieldbackground="white", foreground="black")
    style.configure("Student.Treeview.Heading", font=BUTTON_FONT)

    columns = ("Date", "Time", "Semester", "Subject", "Class", "Faculty")
    tree = ttk.Treeview(frame, columns=columns, show="headings", style="Student.Treeview")
    for col in columns:
        tree.heading(col, text=col)
    tree.column("Date", width=90, anchor="center")
    tree.column("Time", width=80, anchor="center")
    tree.column("Semester", width=110, anchor="center")
    tree.column("Subject", width=200, anchor="w")
    tree.column("Class", width=80, anchor="center")
    tree.column("Faculty", width=160, anchor="w")

    vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=vsb.set)
    tree.grid(row=0, column=0, sticky="nsew")
    vsb.grid(row=0, column=1, sticky="ns")
    frame.grid_rowconfigure(0, weight=1)
    frame.grid_columnconfigure(0, weight=1)

    for _, r in df_all.iterrows():
        tree.insert("", "end", values=(r["Date"], r["Time"], r.get("Semester", ""), r.get("Subject", ""), r.get("Class", ""), r.get("Faculty", "")))


# ===================== DASHBOARD & REPORTS & ML =====================
def open_dashboard_window():
    df_all = load_all_attendance()
    if df_all is None or df_all.empty:
        messagebox.showinfo("Dashboard", "No attendance data available yet.")
        return
    today_str = datetime.now().strftime("%Y-%m-%d")
    df_today = df_all[df_all["Date"] == today_str]

    total_records_today = len(df_today)
    unique_students_today = df_today["Id"].nunique()
    unique_subjects_today = df_today["Subject"].nunique()

    subj_counts = df_today.groupby("Subject")["Id"].nunique().reset_index().sort_values("Id", ascending=False)

    win = tk.Toplevel(root)
    win.title("Attendance Dashboard – Today")
    win.geometry("900x500")
    win.configure(bg=BG_DARK)
    win.grab_set()

    tk.Label(win, text=f"Dashboard – {today_str}", font=SUBTITLE_FONT, bg=BG_DARK, fg=FG_MAIN).pack(pady=8)
    stats_frame = tk.Frame(win, bg=BG_DARK)
    stats_frame.pack(pady=5, fill="x")

    def stat_label(parent, title, value, col):
        box = tk.Frame(parent, bg="#ffffff", bd=1, relief="solid")
        box.grid(row=0, column=col, padx=10, pady=5, sticky="w")
        tk.Label(box, text=title, font=LABEL_FONT, bg="#ffffff").pack(padx=10, pady=4)
        tk.Label(box, text=str(value), font=("Helvetica", 16, "bold"), bg="#ffffff").pack(padx=10, pady=4)

    stat_label(stats_frame, "Total Records Today", total_records_today, 0)
    stat_label(stats_frame, "Unique Students Present", unique_students_today, 1)
    stat_label(stats_frame, "Subjects Conducted Today", unique_subjects_today, 2)

    chart_frame = tk.Frame(win, bg=BG_DARK)
    chart_frame.pack(fill="both", expand=True, padx=10, pady=10)
    if subj_counts.empty:
        tk.Label(chart_frame, text="No subject-wise data for today.", bg=BG_DARK, fg=FG_MAIN, font=LABEL_FONT).pack(pady=20)
        return

    subjects = subj_counts["Subject"].tolist()
    counts = subj_counts["Id"].tolist()

    fig = Figure(figsize=(6, 3), dpi=100)
    ax = fig.add_subplot(111)
    ax.bar(subjects, counts)
    ax.set_xlabel("Subject")
    ax.set_ylabel("No. of Students Present")
    ax.set_title("Subject-wise Attendance Today")
    ax.tick_params(axis="x", rotation=45)

    canvas = FigureCanvasTkAgg(fig, master=chart_frame)
    canvas.draw()
    canvas.get_tk_widget().pack(fill="both", expand=True)


def generate_daily_pdf_report():
    df_all = load_all_attendance()
    if df_all is None or df_all.empty:
        messagebox.showinfo("PDF Report", "No attendance data available.")
        return
    today_str = datetime.now().strftime("%Y-%m-%d")
    df_today = df_all[df_all["Date"] == today_str]
    if df_today.empty:
        messagebox.showinfo("PDF Report", f"No records for today ({today_str}).")
        return
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
    except ImportError:
        messagebox.showerror("Missing Library", "reportlab is not installed.\nInstall it with:\n\npip install reportlab")
        return

    pdf_name = f"Attendance_Report_{today_str}.pdf"
    pdf_path = os.path.join(REPORT_DIR, pdf_name)

    c = canvas.Canvas(pdf_path, pagesize=A4)
    width, height = A4

    c.setFont("Helvetica-Bold", 16)
    c.drawString(50, height - 50, f"Daily Attendance Report - {today_str}")
    c.setFont("Helvetica", 11)
    y = height - 90
    line_height = 14

    grouped = df_today.sort_values(["Semester", "Subject", "Class", "Id"])
    c.drawString(50, y, "Format: Semester | Subject | Class | USN | Name | Time | Faculty")
    y -= 2 * line_height

    for _, r in grouped.iterrows():
        line = (f"{r['Semester']} | {r['Subject']} | {r['Class']} | " f"{r['Id']} | {r['Name']} | {r['Time']} | {r['Faculty']}")
        if y < 80:
            c.showPage()
            c.setFont("Helvetica", 11)
            y = height - 80
        c.drawString(50, y, line)
        y -= line_height
    c.showPage()
    c.save()
    messagebox.showinfo("PDF Report", f"Daily report generated:\n{pdf_name}\n\nSaved in:\n{REPORT_DIR}")


def train_and_show_attendance_risk():
    df_all = load_all_attendance()
    if df_all is None or df_all.empty:
        messagebox.showinfo("Prediction", "No attendance data available.")
        return
    try:
        from sklearn.ensemble import RandomForestClassifier
        import joblib
    except ImportError:
        messagebox.showerror("Missing Library", "scikit-learn or joblib not installed.\nInstall with:\n\npip install scikit-learn joblib")
        return

    df_all["Date_dt"] = pd.to_datetime(df_all["Date"], errors="coerce")
    grp = df_all.groupby(["Id", "Name", "Subject"])
    summary = grp.agg(total_classes=("Date", "count"), present_days=("Date_dt", "nunique"),).reset_index()
    summary["attendance_ratio"] = summary.apply(lambda r: r["present_days"] / r["total_classes"] if r["total_classes"] > 0 else 0.0, axis=1)
    summary["is_defaulter"] = (summary["attendance_ratio"] < 0.75).astype(int)
    if summary["is_defaulter"].nunique() == 1:
        messagebox.showinfo("Prediction", "Data is not diverse enough (all students are same class: all defaulters or all non-defaulters).\nTry again after more data is collected.")
        return
    X = summary[["total_classes", "present_days", "attendance_ratio"]]
    y = summary["is_defaulter"]
    model = RandomForestClassifier(n_estimators=100, random_state=42)
    model.fit(X, y)
    model_path = os.path.join(MODEL_DIR, "attendance_rf_model.joblib")
    import joblib
    joblib.dump(model, model_path)
    summary["predicted_defaulter"] = model.predict(X)

    win = tk.Toplevel(root)
    win.title("Predicted Attendance Risk (RandomForest)")
    win.geometry("800x450")
    win.configure(bg=BG_DARK)
    win.grab_set()
    tk.Label(win, text="Predicted Defaulter Risk – 1 = Likely Defaulter (<75%)", font=SUBTITLE_FONT, bg=BG_DARK, fg=FG_MAIN).pack(pady=8)
    frame = tk.Frame(win, bg=BG_DARK)
    frame.pack(fill="both", expand=True, padx=10, pady=10)

    style = ttk.Style(win)
    style.theme_use("clam")
    style.configure("Predict.Treeview", font=LABEL_FONT, rowheight=24, background="white", fieldbackground="white", foreground="black")
    style.configure("Predict.Treeview.Heading", font=BUTTON_FONT)

    columns = ("Id", "Name", "Subject", "Total", "PresentDays", "Ratio", "Predicted")
    tree = ttk.Treeview(frame, columns=columns, show="headings", style="Predict.Treeview")
    for col in columns:
        tree.heading(col, text=col)
    tree.column("Id", width=60, anchor="center")
    tree.column("Name", width=150, anchor="w")
    tree.column("Subject", width=150, anchor="w")
    tree.column("Total", width=80, anchor="center")
    tree.column("PresentDays", width=100, anchor="center")
    tree.column("Ratio", width=80, anchor="center")
    tree.column("Predicted", width=100, anchor="center")

    vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=vsb.set)
    tree.grid(row=0, column=0, sticky="nsew")
    vsb.grid(row=0, column=1, sticky="ns")
    frame.grid_rowconfigure(0, weight=1)
    frame.grid_columnconfigure(0, weight=1)

    for _, r in summary.iterrows():
        tree.insert("", "end", values=(r["Id"], r["Name"], r["Subject"], int(r["total_classes"]), int(r["present_days"]), f"{r['attendance_ratio']:.2f}", int(r["predicted_defaulter"])))

# ===================== HOD / PRINCIPAL SUPPORT FUNCTIONS =====================
def hod_view_all():
    df_all = load_all_attendance()
    if df_all is None or df_all.empty:
        messagebox.showinfo("Info", "No attendance records found.")
        return
    win = tk.Toplevel(root)
    win.title("Attendance Reports")
    win.geometry("1000x550")
    win.configure(bg=BG_DARK)
    win.grab_set()

    title_text = "HOD – Attendance Reports" if current_role == "hod" else "Principal – Attendance Reports"
    tk.Label(win, text=title_text, font=SUBTITLE_FONT, bg=BG_DARK, fg=FG_MAIN).pack(pady=8)
    filter_frame = tk.Frame(win, bg=BG_DARK)
    filter_frame.pack(fill="x", padx=10, pady=5)

    sem_filter_var = tk.StringVar(value="All")
    subj_filter_var = tk.StringVar(value="All")
    fac_filter_var = tk.StringVar(value="All")
    date_filter_var = tk.StringVar(value="")

    semesters = ["All"] + sorted(set(df_all["Semester"]))
    subjects = ["All"] + sorted(set(df_all["Subject"]))
    faculties = ["All"] + sorted(set(df_all["Faculty"]))

    tk.Label(filter_frame, text="Semester:", bg=BG_DARK, fg=FG_MAIN).grid(row=0, column=0, padx=4)
    sem_cb = ttk.Combobox(filter_frame, textvariable=sem_filter_var, values=semesters, width=15, state="readonly")
    sem_cb.grid(row=0, column=1, padx=4)

    tk.Label(filter_frame, text="Subject:", bg=BG_DARK, fg=FG_MAIN).grid(row=0, column=2, padx=4)
    subj_cb = ttk.Combobox(filter_frame, textvariable=subj_filter_var, values=subjects, width=20, state="readonly")
    subj_cb.grid(row=0, column=3, padx=4)

    tk.Label(filter_frame, text="Faculty:", bg=BG_DARK, fg=FG_MAIN).grid(row=0, column=4, padx=4)
    fac_cb = ttk.Combobox(filter_frame, textvariable=fac_filter_var, values=faculties, width=20, state="readonly")
    fac_cb.grid(row=0, column=5, padx=4)

    tk.Label(filter_frame, text="Date (YYYY-MM-DD):", bg=BG_DARK, fg=FG_MAIN).grid(row=0, column=6, padx=4)
    date_entry = tk.Entry(filter_frame, textvariable=date_filter_var, width=12)
    date_entry.grid(row=0, column=7, padx=4)

    table_frame = tk.Frame(win, bg=BG_DARK)
    table_frame.pack(fill="both", expand=True, padx=10, pady=10)

    style = ttk.Style(win)
    style.theme_use("clam")
    style.configure("Attendance.Treeview", font=LABEL_FONT, rowheight=24, background="white", fieldbackground="white", foreground="black")
    style.configure("Attendance.Treeview.Heading", font=BUTTON_FONT)

    columns = ("Id", "Name", "Class", "Semester", "Subject", "Faculty", "Date", "Time")
    tree = ttk.Treeview(table_frame, columns=columns, show="headings", style="Attendance.Treeview")
    for col in columns:
        tree.heading(col, text=col)

    tree.column("Id", width=120, anchor="center")
    tree.column("Name", width=160, anchor="w")
    tree.column("Class", width=70, anchor="center")
    tree.column("Semester", width=110, anchor="center")
    tree.column("Subject", width=180, anchor="w")
    tree.column("Faculty", width=150, anchor="w")
    tree.column("Date", width=90, anchor="center")
    tree.column("Time", width=80, anchor="center")

    vsb = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=vsb.set)
    tree.grid(row=0, column=0, sticky="nsew")
    vsb.grid(row=0, column=1, sticky="ns")
    table_frame.grid_rowconfigure(0, weight=1)
    table_frame.grid_columnconfigure(0, weight=1)

    def populate_tree(df):
        for row in tree.get_children():
            tree.delete(row)
        for _, r in df.iterrows():
            tree.insert("", "end", values=(r["Id"], r["Name"], r["Class"], r["Semester"], r["Subject"], r["Faculty"], r["Date"], r["Time"]))

    populate_tree(df_all)

    def apply_filters():
        df = df_all.copy()
        if sem_filter_var.get() != "All":
            df = df[df["Semester"] == sem_filter_var.get()]
        if subj_filter_var.get() != "All":
            df = df[df["Subject"] == subj_filter_var.get()]
        if fac_filter_var.get() != "All":
            df = df[df["Faculty"] == fac_filter_var.get()]
        if date_filter_var.get().strip():
            df = df[df["Date"] == date_filter_var.get().strip()]
        populate_tree(df)

    tk.Button(filter_frame, text="Apply Filters", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=apply_filters).grid(row=0, column=8, padx=6)

# ===================== LIVE MONITORING =====================
def open_live_camera():
    if not hasattr(cv2, "VideoCapture"):
        messagebox.showerror("Error", "OpenCV video capture not available.")
        return
    if not os.path.exists(CASCADE_PATH):
        messagebox.showerror("Error", "haarcascade_frontalface_default.xml not found.")
        return
    face_cascade = cv2.CascadeClassifier(CASCADE_PATH)
    cam = cv2.VideoCapture(0)
    if not cam.isOpened():
        messagebox.showerror("Error", "Cannot open camera.")
        return
    set_status("Monitoring... Press Q on the window to close.")
    MAX_ALLOWED = 80
    try:
        while True:
            ret, frame = cam.read()
            if not ret:
                break
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(gray, 1.2, 5)
            count = len(faces)
            if count == 0:
                alert_text = "ALERT: Classroom empty"; alert_color = (0, 0, 255)
            elif count > MAX_ALLOWED:
                alert_text = f"ALERT: Too many people ({count})"; alert_color = (0, 0, 255)
            else:
                alert_text = f"People detected: {count}"; alert_color = (0, 255, 0)
            for (x, y, w, h) in faces:
                cv2.rectangle(frame, (x, y), (x + w, y + h), (255, 255, 255), 2)
            header = f"{current_role.upper() if current_role else 'MONITOR'} LIVE VIEW"
            cv2.putText(frame, header, (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (255, 255, 0), 2)
            cv2.putText(frame, alert_text, (10, 70), cv2.FONT_HERSHEY_SIMPLEX, 0.8, alert_color, 2)
            cv2.imshow("Live Classroom Monitoring – Press Q to close", frame)
            if cv2.waitKey(1) & 0xFF == ord("q"):
                break
    finally:
        cam.release()
        cv2.destroyAllWindows()
        set_status("Live camera view closed.")

# ===================== FACULTY VIEW-ATTENDANCE =====================
def faculty_view_my_attendance():
    global current_faculty_name
    if not current_faculty_name:
        messagebox.showerror("Error", "Faculty not identified. Please login again.")
        return
    df_all = load_all_attendance()
    if df_all is None or df_all.empty:
        messagebox.showinfo("Info", "No attendance records found.")
        return
    df_faculty = df_all[df_all["Faculty"].fillna("") == current_faculty_name]
    if df_faculty.empty:
        messagebox.showinfo("Info", f"No attendance records found for your classes.\n\nLogged in as: {current_faculty_name}\nMake sure you took attendance AFTER using this updated code.")
        return
    win = tk.Toplevel(root)
    win.title(f"My Attendance Records – {current_faculty_name}")
    win.geometry("1000x550")
    win.configure(bg=BG_DARK)
    win.grab_set()
    tk.Label(win, text=f"My Attendance Records – {current_faculty_name}", font=SUBTITLE_FONT, bg=BG_DARK, fg=FG_MAIN).pack(pady=8)
    filter_frame = tk.Frame(win, bg=BG_DARK)
    filter_frame.pack(fill="x", padx=10, pady=5)
    sem_filter_var = tk.StringVar(value="All")
    subj_filter_var = tk.StringVar(value="All")
    date_filter_var = tk.StringVar(value="")
    semesters = ["All"] + sorted(set(df_faculty["Semester"]))
    subjects = ["All"] + sorted(set(df_faculty["Subject"]))
    tk.Label(filter_frame, text="Semester:", bg=BG_DARK, fg=FG_MAIN).grid(row=0, column=0, padx=4)
    sem_cb = ttk.Combobox(filter_frame, textvariable=sem_filter_var, values=semesters, width=15, state="readonly")
    sem_cb.grid(row=0, column=1, padx=4)
    tk.Label(filter_frame, text="Subject:", bg=BG_DARK, fg=FG_MAIN).grid(row=0, column=2, padx=4)
    subj_cb = ttk.Combobox(filter_frame, textvariable=subj_filter_var, values=subjects, width=20, state="readonly")
    subj_cb.grid(row=0, column=3, padx=4)
    tk.Label(filter_frame, text="Date (YYYY-MM-DD):", bg=BG_DARK, fg=FG_MAIN).grid(row=0, column=4, padx=4)
    date_entry = tk.Entry(filter_frame, textvariable=date_filter_var, width=12); date_entry.grid(row=0, column=5, padx=4)

    table_frame = tk.Frame(win, bg=BG_DARK); table_frame.pack(fill="both", expand=True, padx=10, pady=10)
    style = ttk.Style(win); style.theme_use("clam")
    style.configure("FacultyAttendance.Treeview", font=LABEL_FONT, rowheight=24, background="white", fieldbackground="white", foreground="black")
    style.configure("FacultyAttendance.Treeview.Heading", font=BUTTON_FONT)
    columns = ("Id", "Name", "Class", "Semester", "Subject", "Date", "Time")
    tree = ttk.Treeview(table_frame, columns=columns, show="headings", style="FacultyAttendance.Treeview")
    for col in columns:
        tree.heading(col, text=col)
    tree.column("Id", width=120, anchor="center")
    tree.column("Name", width=160, anchor="w")
    tree.column("Class", width=70, anchor="center")
    tree.column("Semester", width=110, anchor="center")
    tree.column("Subject", width=180, anchor="w")
    tree.column("Date", width=90, anchor="center")
    tree.column("Time", width=80, anchor="center")
    vsb = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview); tree.configure(yscrollcommand=vsb.set)
    tree.grid(row=0, column=0, sticky="nsew"); vsb.grid(row=0, column=1, sticky="ns")
    table_frame.grid_rowconfigure(0, weight=1); table_frame.grid_columnconfigure(0, weight=1)

    def populate_tree(df):
        for row in tree.get_children():
            tree.delete(row)
        for _, r in df.iterrows():
            tree.insert("", "end", values=(r["Id"], r["Name"], r["Class"], r["Semester"], r["Subject"], r["Date"], r["Time"]))
    populate_tree(df_faculty)

    def apply_filters():
        df = df_faculty.copy()
        if sem_filter_var.get() != "All":
            df = df[df["Semester"] == sem_filter_var.get()]
        if subj_filter_var.get() != "All":
            df = df[df["Subject"] == subj_filter_var.get()]
        if date_filter_var.get().strip():
            df = df[df["Date"] == date_filter_var.get().strip()]
        populate_tree(df)
    tk.Button(filter_frame, text="Apply Filters", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=apply_filters).grid(row=0, column=6, padx=6)

# ===================== CHATBOT (local) - integrates docx answers =====================
def process_chat_query(query: str):
    q = query.lower().strip()
    # try doc answer first
    doc_answer = answer_from_doc(q) if DEPT_SENTENCES else None
    if doc_answer:
        return doc_answer
    m = re.search(r"is\s+([a-z\s]+)\s+present", q)
    if m:
        name = m.group(1).strip()
        matches = check_name_present_today(name)
        if matches:
            first = matches[0]
            return f"Yes — {first['Name']} is present (marked at {first['Time']})."
        else:
            return f"No matching attendance record found for '{name}' today."
    if "how many" in q and "present" in q:
        df_all = load_all_attendance()
        today = datetime.now().strftime("%Y-%m-%d")
        if df_all is None:
            return "No attendance data available."
        df_today = df_all[df_all["Date"] == today]
        n = df_today["Id"].nunique()
        return f"{n} students are present today."
    if "who is present" in q or "who's present" in q:
        df_all = load_all_attendance()
        today = datetime.now().strftime("%Y-%m-%d")
        if df_all is None:
            return "No attendance data available."
        df_today = df_all[df_all["Date"] == today]
        names = df_today["Name"].unique().tolist()
        if not names:
            return "No one is marked present yet."
        to_show = names[:12]
        s = ", ".join(to_show)
        if len(names) > 12:
            s += f", and {len(names)-12} more."
        return f"Present today: {s}"
    return "I can answer attendance queries or departmental questions from the guide (try: 'Who is HOD?', 'How many labs?', 'Tell me achievements')."

def open_chatbot_window():
    win = tk.Toplevel(root)
    win.title("Attendance Chatbot")
    win.geometry("650x520")
    win.configure(bg=BG_DARK)
    win.grab_set()

    tk.Label(win, text="Attendance Chatbot (local)", font=SUBTITLE_FONT, bg=BG_DARK, fg=FG_MAIN).pack(pady=8)
    chat_frame = tk.Frame(win, bg="#ffffff", bd=1, relief="solid")
    chat_frame.pack(fill="both", expand=True, padx=12, pady=8)
    text_area = tk.Text(chat_frame, state="disabled", wrap="word", bg="white", fg="black", font=("Helvetica", 11))
    text_area.pack(fill="both", expand=True, padx=6, pady=6)

    entry_var = tk.StringVar()
    entry = tk.Entry(win, textvariable=entry_var, font=LABEL_FONT, bg=ENTRY_BG, fg=ENTRY_FG)
    entry.pack(fill="x", padx=12, pady=6)

    def append_chat(who, msg):
        text_area.configure(state="normal")
        text_area.insert("end", f"{who}: {msg}\n\n")
        text_area.see("end")
        text_area.configure(state="disabled")

    def send_text():
        q = entry_var.get().strip()
        if not q:
            return
        append_chat("You", q)
        entry_var.set("")
        reply = process_chat_query(q)
        append_chat("Alan", reply)
        speak_text(reply)

    # Voice controls (Start/Stop recording)
    controls_frame = tk.Frame(win, bg=BG_DARK)
    controls_frame.pack(pady=6)

    btn_start = tk.Button(controls_frame, text="Start Query (Record)", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT,
                          command=start_manual_recording, width=20)
    btn_start.grid(row=0, column=0, padx=6)

    btn_stop = tk.Button(controls_frame, text="Stop Query (Process)", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT,
                         command=stop_manual_recording_and_process, width=20)
    btn_stop.grid(row=0, column=1, padx=6)

    def voice_query_once():
        r = sr.Recognizer()
        try:
            with sr.Microphone() as src:
                set_status("Listening for chatbot question (short)...")
                r.adjust_for_ambient_noise(src, duration=0.5)
                audio = r.listen(src, phrase_time_limit=12)
            try:
                q = r.recognize_google(audio)
            except Exception:
                q = ""
            if not q:
                append_chat("You (voice)", "(unrecognized)")
                append_chat("Alan", "Sorry — couldn't understand the voice question.")
                return
            append_chat("You (voice)", q)
            reply = process_chat_query(q)
            append_chat("Alan", reply)
            speak_text(reply)
        except Exception as e:
            append_chat("Alan", "Voice input failed.")
            set_status("Chatbot voice failed. Check mic.")

    tk.Button(win, text="Voice Question (single-shot)", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=voice_query_once).pack(pady=4)
    tk.Button(win, text="Send", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=send_text).pack(pady=6)

    # show basic hint
    append_chat("Alan", "Hi — say 'Hey Alan' to wake the voice assistant, or use Start/Stop Query to record a question and get an attendance answer (e.g., 'Is Sunanda present today?').")

# ===================== AUTH UI & BUILD FRAMES =====================
def authorize(role, success_callback):
    global current_role, current_faculty_name, current_faculty_sem_subjects, current_faculty_sections
    login_win = tk.Toplevel(root)
    login_win.title(f"{role.capitalize()} Login")
    login_win.geometry("350x220")
    login_win.configure(bg=BG_DARK)
    login_win.grab_set()

    tk.Label(login_win, text=f"{role.capitalize()} Login", font=SUBTITLE_FONT, bg=BG_DARK, fg=FG_MAIN).pack(pady=12)
    tk.Label(login_win, text="Username:", bg=BG_DARK, fg=FG_MAIN).pack()
    user_entry = tk.Entry(login_win, font=LABEL_FONT, bg=ENTRY_BG, fg=ENTRY_FG); user_entry.pack(pady=4)
    tk.Label(login_win, text="Password:", bg=BG_DARK, fg=FG_MAIN).pack()
    pass_entry = tk.Entry(login_win, show="*", font=LABEL_FONT, bg=ENTRY_BG, fg=ENTRY_FG); pass_entry.pack(pady=4)

    def attempt_login():
        global current_role, current_faculty_name, current_faculty_sem_subjects, current_faculty_sections
        u = user_entry.get().strip(); p = pass_entry.get().strip()
        if role == "faculty":
            fac_users = AUTH["faculty"]
            if u in fac_users and p == fac_users[u]["password"]:
                fk = fac_users[u]["faculty_key"]
                info = FACULTY_INFO.get(fk)
                if info is None:
                    messagebox.showerror("Error", "Faculty mapping not found. Check FACULTY_INFO."); return
                current_role = "faculty"
                current_faculty_name = info["display_name"]
                current_faculty_sem_subjects = info.get("sem_subjects", {})
                current_faculty_sections = info.get("sections", {})
                messagebox.showinfo("Success", f"Welcome, {current_faculty_name}")
                login_win.destroy(); success_callback(); return
        elif role == "hod":
            hod_users = AUTH["hod"]
            if u in hod_users and p == hod_users[u]["password"]:
                current_role = "hod"; current_faculty_name = hod_users[u]["name"]; current_faculty_sem_subjects = {}; current_faculty_sections = {}
                messagebox.showinfo("Success", "HOD Login Successful"); login_win.destroy(); success_callback(); return
        elif role == "principal":
            pri_users = AUTH["principal"]
            if u in pri_users and p == pri_users[u]["password"]:
                current_role = "principal"; current_faculty_name = pri_users[u]["name"]; current_faculty_sem_subjects = {}; current_faculty_sections = {}
                messagebox.showinfo("Success", "Principal Login Successful"); login_win.destroy(); success_callback(); return
        messagebox.showerror("Access Denied", "Invalid username or password")
    tk.Button(login_win, text="Login", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=attempt_login).pack(pady=15)

def make_entry(parent):
    e = tk.Entry(parent, font=LABEL_FONT, bg=ENTRY_BG, fg=ENTRY_FG)
    return e

def create_role_button(parent, text, command):
    btn = tk.Button(parent, text=text, font=BUTTON_FONT, bg=BUTTON_BG, fg=BUTTON_FG, activebackground=BUTTON_ACTIVE_BG, activeforeground=BUTTON_FG, width=16, height=2, command=command)
    return btn

def build_role_frame():
    frame = tk.Frame(root, bg=BG_DARK)
    frame.grid(row=0, column=0, sticky="nsew")
    frames["role"] = frame
    for i in range(3):
        frame.grid_rowconfigure(i, weight=1)
    frame.grid_columnconfigure(0, weight=1)

    title = tk.Label(frame, text="Face Recognition Based Attendance System", font=TITLE_FONT, bg=BG_DARK, fg=FG_MAIN)
    title.pack(pady=18)

    subtitle = tk.Label(frame, text="Select your role", font=SUBTITLE_FONT, bg=BG_DARK, fg=FG_MAIN)
    subtitle.pack(pady=6)

    btn_frame = tk.Frame(frame, bg=BG_DARK)
    btn_frame.pack(pady=10)

    b1 = create_role_button(btn_frame, "Student Login", lambda: show_frame("student"))
    b2 = create_role_button(btn_frame, "Faculty Login", lambda: authorize("faculty", lambda: show_frame("faculty")))
    b3 = create_role_button(btn_frame, "HOD Login", lambda: authorize("hod", lambda: show_frame("hod")))
    b4 = create_role_button(btn_frame, "Principal Login", lambda: authorize("principal", lambda: show_frame("principal")))

    b1.grid(row=0, column=0, padx=14)
    b2.grid(row=0, column=1, padx=14)
    b3.grid(row=0, column=2, padx=14)
    b4.grid(row=0, column=3, padx=14)

    tk.Button(frame, text="Open Attendance Chatbot", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=open_chatbot_window).pack(pady=8)

def build_student_frame():
    global student_usn_entry
    frame = tk.Frame(root, bg=BG_DARK); frame.grid(row=0, column=0, sticky="nsew"); frames["student"] = frame
    tk.Label(frame, text="Student – View Attendance", font=TITLE_FONT, bg=BG_DARK, fg=FG_MAIN).pack(pady=20)
    content = tk.Frame(frame, bg=BG_DARK); content.pack(pady=20)
    tk.Label(content, text="USN:", bg=BG_DARK, fg=FG_MAIN, font=LABEL_FONT).grid(row=0, column=0, padx=10, pady=5, sticky="e")
    student_usn_entry = tk.Entry(content, font=LABEL_FONT, bg=ENTRY_BG, fg=ENTRY_FG); student_usn_entry.grid(row=0, column=1, padx=10, pady=5)
    btn_today = tk.Button(content, text="Check Today's Attendance", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=student_check_today, width=24); btn_today.grid(row=1, column=0, columnspan=2, pady=10)
    btn_hist = tk.Button(content, text="View Full History", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=student_view_history, width=24); btn_hist.grid(row=2, column=0, columnspan=2, pady=10)
    tk.Button(frame, text="Back", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=lambda: show_frame("role")).pack(pady=10)

def build_faculty_frame():
    global entry_usn, entry_name, entry_class, sem_var, subject_var, subject_menu
    frame = tk.Frame(root, bg=BG_DARK); frame.grid(row=0, column=0, sticky="nsew"); frames["faculty"] = frame
    tk.Label(frame, text="Faculty – Attendance Panel", font=TITLE_FONT, bg=BG_DARK, fg=FG_MAIN).pack(pady=12)
    main = tk.Frame(frame, bg=BG_DARK); main.pack(fill="both", expand=True, padx=20, pady=10)
    left = tk.Frame(main, bg=BG_DARK); right = tk.Frame(main, bg=BG_DARK)
    left.pack(side="left", fill="y", padx=10); right.pack(side="right", fill="both", expand=True, padx=10)

    tk.Label(left, text="Register Student & Capture Images", font=SUBTITLE_FONT, bg=BG_DARK, fg=FG_MAIN).pack(pady=6)
    tk.Label(left, text="USN (e.g. 3PD22AI900):", bg=BG_DARK, fg=FG_MAIN, font=LABEL_FONT).pack(anchor="w")
    entry_usn = make_entry(left); entry_usn.pack(pady=4, fill="x")
    tk.Label(left, text="Student Name:", bg=BG_DARK, fg=FG_MAIN, font=LABEL_FONT).pack(anchor="w")
    entry_name = make_entry(left); entry_name.pack(pady=4, fill="x")
    tk.Label(left, text="Semester:", bg=BG_DARK, fg=FG_MAIN, font=LABEL_FONT).pack(anchor="w")
    entry_class = make_entry(left); entry_class.pack(pady=4, fill="x")

    tk.Button(left, text="Save & Capture Images", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=take_images, wraplength=180).pack(pady=8, fill="x")
    tk.Button(left, text="Train Model", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=train_images).pack(pady=6, fill="x")
    tk.Button(left, text="Evaluate Trained Model", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=evaluate_trained_model).pack(pady=6, fill="x")
    tk.Button(left, text="Remove Student (USN)", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=lambda: remove_student_by_usn(entry_usn.get().strip())).pack(pady=6, fill="x")
    tk.Button(left, text="Reset All Trained Students", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=lambda: reset_all_trained_students()).pack(pady=6, fill="x")
    tk.Button(left, text="Back", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=lambda: show_frame("role")).pack(pady=12, fill="x")

    tk.Label(right, text="Take Attendance (Face Recognition)", font=SUBTITLE_FONT, bg=BG_DARK, fg=FG_MAIN).pack(pady=8)
    controls = tk.Frame(right, bg=BG_DARK); controls.pack(pady=6, fill="x")
    tk.Label(controls, text="Semester:", bg=BG_DARK, fg=FG_MAIN, font=LABEL_FONT).grid(row=0, column=0, padx=8, pady=4, sticky="w")
    sem_var = tk.StringVar(value=SEMESTER_OPTIONS[0]); subject_var = tk.StringVar()

    def update_subjects_for_semester(*args):
        sem = sem_var.get(); subjects = []
        if current_faculty_sem_subjects and current_role == "faculty":
            subjects = current_faculty_sem_subjects.get(sem, [])
        if not subjects:
            subjects = SUBJECTS_BY_SEM.get(sem, [])
        if not subjects:
            subjects = ["(No subjects mapped)"]
        subject_var.set(subjects[0])
        subject_menu["menu"].delete(0, "end")
        for subj in subjects:
            subject_menu["menu"].add_command(label=subj, command=lambda s=subj: subject_var.set(s))
        if entry_class is not None and current_faculty_sections and current_role == "faculty":
            sec = current_faculty_sections.get(sem, "")
            if sec:
                entry_class.delete(0, tk.END); entry_class.insert(0, sec)

    sem_menu = tk.OptionMenu(controls, sem_var, *SEMESTER_OPTIONS, command=lambda _: update_subjects_for_semester())
    sem_menu.config(font=LABEL_FONT, bg=BUTTON_BG, fg=BUTTON_FG); sem_menu.grid(row=0, column=1, padx=8, pady=4, sticky="w")
    tk.Label(controls, text="Subject:", bg=BG_DARK, fg=FG_MAIN, font=LABEL_FONT).grid(row=1, column=0, padx=8, pady=4, sticky="w")
    subject_menu = tk.OptionMenu(controls, subject_var, ""); subject_menu.config(font=LABEL_FONT, bg=BUTTON_BG, fg=BUTTON_FG); subject_menu.grid(row=1, column=1, padx=8, pady=4, sticky="w")
    update_subjects_for_semester()

    tk.Button(right, text="Start Attendance (Recognize Faces)", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=track_images, width=28).pack(pady=12)
    tk.Button(right, text="View My Attendance Records", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=faculty_view_my_attendance, width=28).pack(pady=6)
    tk.Button(right, text="Open Dashboard (Today)", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=open_dashboard_window, width=28).pack(pady=6)
    tk.Button(right, text="Attendance Chatbot", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=open_chatbot_window, width=20).pack(pady=6)

def build_hod_frame():
    frame = tk.Frame(root, bg=BG_DARK); frame.grid(row=0, column=0, sticky="nsew"); frames["hod"] = frame
    tk.Label(frame, text="HOD – Attendance & Monitoring", font=TITLE_FONT, bg=BG_DARK, fg=FG_MAIN).pack(pady=20)
    tk.Button(frame, text="View All Attendance (with Filters)", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=hod_view_all, width=30).pack(pady=10)
    tk.Button(frame, text="Open Dashboard (Today)", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=open_dashboard_window, width=30).pack(pady=10)
    tk.Button(frame, text="Generate Today's PDF Report", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=generate_daily_pdf_report, width=30).pack(pady=10)
    tk.Button(frame, text="Predict Defaulters (ML)", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=train_and_show_attendance_risk, width=30).pack(pady=10)
    tk.Button(frame, text="Open Live Camera", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=open_live_camera, width=30).pack(pady=10)
    tk.Button(frame, text="Back", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=lambda: show_frame("role"), width=15).pack(pady=20)

def build_principal_frame():
    frame = tk.Frame(root, bg=BG_DARK); frame.grid(row=0, column=0, sticky="nsew"); frames["principal"] = frame
    tk.Label(frame, text="Principal – Attendance & Monitoring", font=TITLE_FONT, bg=BG_DARK, fg=FG_MAIN).pack(pady=20)
    tk.Button(frame, text="View All Attendance (with Filters)", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=hod_view_all, width=30).pack(pady=10)
    tk.Button(frame, text="Open Dashboard (Today)", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=open_dashboard_window, width=30).pack(pady=10)
    tk.Button(frame, text="Generate Today's PDF Report", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=generate_daily_pdf_report, width=30).pack(pady=10)
    tk.Button(frame, text="Predict Defaulters (ML)", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=train_and_show_attendance_risk, width=30).pack(pady=10)
    tk.Button(frame, text="Open Live Camera", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=open_live_camera, width=30).pack(pady=10)
    tk.Button(frame, text="Back", bg=BUTTON_BG, fg=BUTTON_FG, font=BUTTON_FONT, command=lambda: show_frame("role"), width=15).pack(pady=20)

def show_frame(name: str):
    frame = frames.get(name)
    if frame:
        frame.tkraise()
        set_status("Ready.")

# ===================== RESET ALL TRAINED STUDENTS (GUI helper) =====================
def reset_all_trained_students():
    """GUI-triggered reset — deletes training images, trainer, and resets student CSV."""
    if not messagebox.askyesno("Confirm Reset", "This will DELETE all training images, Trainer.yml and reset StudentDetails.csv. This cannot be undone. Continue?"):
        return
    # delete training images
    deleted = 0
    if os.path.exists(TRAINING_IMAGE_DIR):
        for fname in os.listdir(TRAINING_IMAGE_DIR):
            try:
                os.remove(os.path.join(TRAINING_IMAGE_DIR, fname))
                deleted += 1
            except Exception:
                pass
    # delete trainer
    try:
        if os.path.exists(TRAINER_PATH):
            os.remove(TRAINER_PATH)
    except Exception:
        pass
    # reset students CSV
    try:
        import pandas as pd
        df = pd.DataFrame(columns=["Label","USN","Name","Semester"])
        df.to_csv(STUDENT_CSV, index=False)
    except Exception:
        # fallback - remove file then create minimal csv
        try:
            if os.path.exists(STUDENT_CSV):
                os.remove(STUDENT_CSV)
        except Exception:
            pass
        with open(STUDENT_CSV, "w", encoding="utf-8") as f:
            f.write("Label,USN,Name,Semester\n")
    messagebox.showinfo("Reset Done", f"Deleted {deleted} training images.\nTrainer removed (if present).\nStudentDetails.csv reset.")
    set_status("All trained students removed. Please re-capture faces and retrain the model.")

# ===================== STARTUP =====================
build_role_frame()
build_student_frame()
build_faculty_frame()
build_hod_frame()
build_principal_frame()

show_frame("role")

# speak greeting once the GUI is visible (non-blocking)
def _startup_greeting():
    # non-blocking: will not freeze GUI
    greeting = "Hi — I am Alan. I am an attendance bot made by AIML students."
    speak_text(greeting)
# small delay to let the window render
root.after(600, _startup_greeting)

# start voice listener (background)
start_voice_listener()

# ===================== CLEAN SHUTDOWN =====================
def on_closing():
    try:
        stop_voice_listener()
    except Exception as e:
        print("Failed to stop speech listener:", e)

    try:
        if _tts_engine:
            _tts_engine.stop()
    except Exception as e:
        print("Failed to stop TTS engine:", e)

    try:
        root.destroy()
    except Exception:
        os._exit(0)

root.protocol("WM_DELETE_WINDOW", on_closing)

root.mainloop()
